from __future__ import annotations

import difflib, json, re, unicodedata
from collections import defaultdict
from pathlib import Path
from urllib.parse import urlparse
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
CATALOG = ROOT / "src/data/catalog.json"
OLD_BOOK = ROOT / "Excel WASSAL 2025.xlsx"
NEW_BOOK = ROOT / "alsamah 18-08-2026.xlsx"
MANIFEST = ROOT / "src/data/pending-generated-images.json"

COLOR_FIX = {"blanc cass�":"blanc cassé","multicolores�":"multicolore","soman":"saumon","lanc":"blanc","none":"non renseignée","vert d'eau  ":"vert d'eau"}
HEX = {"noir":"#1d1d1c","blanc":"#f5f5f0","blanc cassé":"#e9e2d2","beige":"#cfb491","gris":"#a8aaad","bleu marine":"#17233d","bleu":"#36608d","bleu ciel":"#91b9d0","rose":"#c9828e","rouge":"#a7473e","orange":"#c8733d","jaune":"#d0a53a","marron":"#6f4a35","violet":"#765d87","vert":"#627b54","bordeaux":"#6e2639","saumon":"#d79079","multicolore":"#8a756f"}
AR = {"noir":"أسود","blanc":"أبيض","blanc cassé":"أوف وايت","beige":"بيج","gris":"رمادي","bleu marine":"كحلي","bleu":"أزرق","bleu ciel":"أزرق سماوي","rose":"وردي","rouge":"أحمر","orange":"برتقالي","jaune":"أصفر","marron":"بني","violet":"بنفسجي","vert":"أخضر","bordeaux":"عنابي","saumon":"سلموني","multicolore":"متعدد الألوان"}
EN = {"noir":"Black","blanc":"White","blanc cassé":"Off-white","beige":"Beige","gris":"Gray","bleu marine":"Navy","bleu":"Blue","bleu ciel":"Sky blue","rose":"Pink","rouge":"Red","orange":"Orange","jaune":"Yellow","marron":"Brown","violet":"Purple","vert":"Green","bordeaux":"Burgundy","saumon":"Salmon","multicolore":"Multicolor"}

def fold(value):
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()

def slug(value): return re.sub(r"[^a-z0-9]+", "-", fold(value)).strip("-") or "produit"

def clean_name(value):
    text = str(value or "").replace("�", "").strip()
    text = re.sub(r"^\d+-", "", text)
    text = re.sub(r"^\d+\s*", "", text)
    text = re.sub(r"\s*-\s*(?:XS|S|M|L|XL|XXL|2XL|3XL|4XL|5XL|6XL|S/M|M/L|L/XL|taille Libre|\d+(?:[,.]\d+)?XL?|\d+)\s*-.*$", "", text, flags=re.I)
    return re.sub(r"\s+", " ", text).strip(" -")

def color(value):
    c = re.sub(r"\s+", " ", str(value or "non renseignée").strip().lower())
    aliases={"blanche":"blanc","blanc casse":"blanc cassé","blanc cass":"blanc cassé","noire":"noir","navy":"bleu marine","grey":"gris","gray":"gris","rose poudr":"rose","vert d'eau":"vert","peau":"beige"}
    return aliases.get(COLOR_FIX.get(c,c), COLOR_FIX.get(c,c))

def category(name):
    n=fold(name)
    if any(x in n for x in ("garcon","boy","fille","girl")): return "enfants"
    if any(x in n for x in ("men ","homme")): return "homme"
    if any(x in n for x in ("socks","chaussette","collant","tight")): return "autres"
    return "femme"

catalog=json.loads(CATALOG.read_text(encoding="utf-8"))
FORCE_NAMES={fold(name) for name in [
    "Pyjama 2 pièces en maille avec motif cœur carreaux et effet bimatière",
    "Slip Hi-Cuts",
    "Pyjama 2 Pièces LOVE - Good Vibes Only à Carreaux Verts",
    "Pyjama 2 Pièces Homewear Bleu Rayé",
    "Pyjama 2 pièces fluide à manches longues",
    "Pyjama 2 pièces Soft Check",
    "Pyjama 2 pièces fluide à manches longues avec bord-côtes contrastants",
    "Pyjama 2 Pièces SWEET TODAY",
    "Pyjama 2 pièces Pure Softness",
    "Pyjama 2 Pièces Homewear",
    "Ensemble Pyjama Quadrillé Everything is Possible",
    "Pyjama à Capuche Angel Baby",
    "Pyjama 2 Pièces LOVE & Pantalon à Carreaux Multicolores",
    "Ensemble Homewear à Capuche Believe",
    "Ensemble Pyjamar Vert d'Eau à Micro-Motifs Fleuris",
    "Ensemble Pyjama 2 Pièces à Carreaux / Pochette Assortie",
    "Ensemble Pyjama Choco / Prune Imprimé Cœurs Love For All",
    "Pyjama femme 2 pièces ALL DAY",
    "Ensemble Pyjama Marines Crew 84",
    "Ensemble Pyjama à Capuche Hello Winter",
    "Ensemble Pyjama à Capuche & Motifs Cœurs",
    "Ensemble Pyjama Friday",
    "Pyjama femme 2 pièces",
    "Ensemble Pyjama Quadrillé",
    "Ensemble Pyjama à Capuche Cool",
    "Ensemble Pyjama Vert Sauge Imprimé Étoiles",
    "Ensemble Pyjama Tête d'Ours",
    "Ensemble Pyjama 2 Pièces à Poche Fleurie",
    "Ensemble Pyjama 2 Pièces à Carreaux",
    "Ensemble Pyjama Ours Hipster",
    "100% Cotton Girls' Tank Top",
]}
old=openpyxl.load_workbook(OLD_BOOK,read_only=True,data_only=True)["bulk_import"]
rows=old.iter_rows(values_only=True); headers=[str(x).strip() if x else "" for x in next(rows)]
old_barcodes={str(dict(zip(headers,row)).get("barcode") or "").strip() for row in rows}
sheet=openpyxl.load_workbook(NEW_BOOK,read_only=True,data_only=True).active
rows=sheet.iter_rows(values_only=True); headers=[str(x).strip() if x else "" for x in next(rows)]
groups=defaultdict(list)
for row in rows:
    item=dict(zip(headers,row)); barcode=str(item.get("Code Barre") or "").strip()
    identity=fold(clean_name(item.get("DESCRIPTION")))
    forced_identity=any(difflib.SequenceMatcher(None,identity,name).ratio() >= .90 for name in FORCE_NAMES)
    if not barcode or (barcode in old_barcodes and not forced_identity): continue
    groups[identity].append(item)

keys=[]
for product in catalog:
    handle=urlparse(product.get("sourcePage") or "").path.rstrip("/").split("/")[-1]
    keys.extend((fold(product["name"]["fr"]),fold(product["name"]["en"]),fold(product["id"]),fold(handle)))

added=[]; pending=[]; used={p["id"] for p in catalog}
for identity, items in groups.items():
    url=next((str(x.get("MODELE")) for x in items if x.get("MODELE")),None)
    handle=fold(urlparse(url or "").path.rstrip("/").split("/")[-1])
    score=max(difflib.SequenceMatcher(None,identity,key).ratio() for key in keys if key)
    if handle: score=max(score,max(difflib.SequenceMatcher(None,handle,key).ratio() for key in keys if key))
    forced_identity=any(difflib.SequenceMatcher(None,identity,name).ratio() >= .90 for name in FORCE_NAMES)
    if score >= .88 and not forced_identity: continue
    name=clean_name(items[0].get("DESCRIPTION"))
    if fold(name) == fold("NIPPLE COVER PAD"): continue
    pid=slug(name); base=pid; i=2
    while pid in used: pid=f"{base}-{i}"; i+=1
    used.add(pid)
    colors=[]
    for cname in dict.fromkeys(color(x.get("Color")) for x in items):
        cid=slug(cname); path=f"/assets/generated/colors/{pid}-{cid}.png"
        colors.append({"id":cid,"label":{"fr":cname.capitalize(),"ar":AR.get(cname,cname),"en":EN.get(cname,cname.capitalize())},"hex":HEX.get(cname,"#8a8882"),"image":None,"imageKind":"missing","lifestyleImage":None})
        pending.append({"productId":pid,"productName":name,"colorId":cid,"color":EN.get(cname,cname),"hex":HEX.get(cname,"#8a8882"),"path":path})
    prices=sorted(float(x["PRIX DE VENTE"]) for x in items if x.get("PRIX DE VENTE") not in (None,"")); regular=prices[len(prices)//2] if prices else None
    stock=sum(float(x.get("QUANTITE") or 0) for x in items); cat=category(name)
    product={"id":pid,"brand":"ALSAMAH","category":cat,"categoryName":{"femme":"Fashion Femme","homme":"Homme","enfants":"Enfants","autres":"Hosiery & Accessories"}[cat],"subcategory":None,"name":{"fr":name,"ar":name,"en":name},"short":{"fr":f"{name}, sélection ALSAMAH.","ar":name,"en":f"{name}, an ALSAMAH essential."},"description":{"fr":f"Produit ALSAMAH disponible dans les tailles et couleurs indiquées.","ar":name,"en":"ALSAMAH product available in the listed sizes and colors."},"regularPrice":regular,"price":round(regular*.7) if regular else None,"stock":round(stock,2),"sizes":sorted({str(x.get("Size") or "Taille unique").strip() for x in items}),"colors":colors,"purchasable":bool(regular and stock>0),"imageStatus":"missing","sourcePage":url,"variantCount":len(items),"missing":{"price":regular is None,"category":False,"image":True}}
    catalog.append(product); added.append(pid)

for product in catalog:
    for variant in product["colors"]:
        image=variant.get("image")
        if image and image.startswith("/assets/") and not (ROOT / "public" / image.lstrip("/")).exists():
            variant["image"]=None; variant["imageKind"]="missing"
    available=any(variant.get("image") for variant in product["colors"])
    if not available:
        product["imageStatus"]="missing"; product["missing"]["image"]=True
CATALOG.write_text(json.dumps(catalog,ensure_ascii=False,indent=2),encoding="utf-8")
MANIFEST.write_text(json.dumps(pending,ensure_ascii=False,indent=2),encoding="utf-8")
print(json.dumps({"added":len(added),"ids":added,"images":len(pending)},ensure_ascii=False))
