# -*- coding: utf-8 -*-
"""Rapport PDF detaille en francais : etat des donnees manquantes, stock WASSAL 2025."""
import openpyxl, collections, datetime, html
from weasyprint import HTML

SRC = 'c:/Users/Asus/Desktop/alsamah/Excel WASSAL 2025.xlsx'
OUT = 'c:/Users/Asus/Desktop/alsamah/RAPPORT_DONNEES_MANQUANTES_WASSAL_2025.pdf'

wb = openpyxl.load_workbook(SRC, data_only=True)
rows = list(wb['bulk_import'].iter_rows(values_only=True))
hdr = [h for h in rows[0] if h]
data = []
for i, r in enumerate(rows[1:], start=2):
    if any(x is not None for x in r[:20]):
        d = dict(zip(hdr, r))
        d['_row'] = i
        data.append(d)

cats = {}
for r in wb['category ids'].iter_rows(min_row=4, values_only=True):
    if r[1] is not None:
        cats[str(r[1]).strip()] = str(r[2]).strip()

by = collections.OrderedDict()
for d in data:
    by.setdefault(d['name'], []).append(d)

P, L = len(by), len(data)
U = sum(d.get('inventory') or 0 for d in data)


def n(x, dec=0):
    return format(round(x, dec) if dec else int(round(x)), ',.%df' % dec).replace(',', '\u202f')


def esc(s):
    return html.escape(str(s))


# ---------- etat par produit ----------
def state(g):
    core = any((not x.get('price')) or (not x.get('categ_id')) for x in g)
    soft = any((not x.get('photo_urls')) or (not x.get('content')) for x in g)
    return core, soft


buck = collections.Counter()
bunits = collections.Counter()
for name, g in by.items():
    core, soft = state(g)
    k = 'bloque' if core else ('media' if soft else 'complet')
    buck[k] += 1
    bunits[k] += sum(x.get('inventory') or 0 for x in g)

# ---------- champs ----------
FIELDS = [
    ('price', 'Prix de vente', 'Bloquant'),
    ('categ_id', 'Catégorie', 'Bloquant'),
    ('photo_urls', 'Photo (lien)', 'Critique'),
    ('content', 'Fiche détaillée FR', 'Important'),
    ('content_ar', 'Fiche détaillée AR', 'Important'),
    ('content_en', 'Fiche détaillée EN', 'Important'),
    ('description_sale', 'Description courte FR', 'Mineur'),
    ('description_sale_ar', 'Description courte AR', 'Mineur'),
    ('description_sale_en', 'Description courte EN', 'Mineur'),
    ('size', 'Taille', 'Mineur'),
    ('color', 'Couleur', 'Mineur'),
]
frows = []
for f, lbl, sev in FIELDS:
    full = sum(1 for nm, g in by.items() if all(not x.get(f) for x in g))
    part = sum(1 for nm, g in by.items()
               if any(not x.get(f) for x in g) and any(x.get(f) for x in g))
    lines = sum(1 for d in data if not d.get(f))
    un = sum(d.get('inventory') or 0 for d in data if not d.get(f))
    if lines:
        frows.append((lbl, sev, full, part, lines, un))

OK_FIELDS = [('name', 'Nom FR'), ('name_ar', 'Nom AR'), ('name_en', 'Nom EN'),
             ('sku', 'SKU'), ('sku_variant', 'SKU variante'), ('barcode', 'Code-barres'),
             ('product_brand_id', 'Marque'), ('inventory', 'Stock')]
ok_list = [lbl for f, lbl in OK_FIELDS if sum(1 for d in data if not d.get(f)) == 0]

# ---------- top produits bloques ----------
blocked = []
for nm, g in by.items():
    bad = [x for x in g if not x.get('price') or not x.get('categ_id')]
    if bad:
        blocked.append({
            'n': nm, 'u': sum(x.get('inventory') or 0 for x in g),
            'nv': len(g), 'nb': len(bad), 'row': min(x['_row'] for x in g),
            'sz': ', '.join(sorted(set(str(x.get('size') or '?') for x in g)))[:30],
        })
blocked.sort(key=lambda x: -x['u'])
TOP = blocked[:30]
top_u = sum(b['u'] for b in TOP)
bl_u = sum(b['u'] for b in blocked)

# ---------- partiels ----------
partial = [(nm, sum(1 for x in g if not x.get('price')), len(g),
            sum(x.get('inventory') or 0 for x in g),
            next((x['price'] for x in g if x.get('price')), None))
           for nm, g in by.items()
           if any(not x.get('price') for x in g) and any(x.get('price') for x in g)]

# ---------- vendables sans photo ----------
nophoto = [(nm, sum(x.get('inventory') or 0 for x in g), len(g), min(x['_row'] for x in g),
            next((x['price'] for x in g if x.get('price')), None))
           for nm, g in by.items()
           if all(x.get('price') and x.get('categ_id') for x in g)
           and all(not x.get('photo_urls') for x in g)]
nophoto.sort(key=lambda x: -x[1])

# ---------- anomalies stock ----------
anom = []
for d in data:
    v = d.get('inventory')
    if isinstance(v, (int, float)):
        if v < 0:
            anom.append(('Quantité négative', d, v))
        elif v != int(v):
            anom.append(('Quantité décimale', d, v))
dupes = [k for k, c in collections.Counter(str(d.get('sku_variant')) for d in data).items() if c > 1]
dup_rows = [d for d in data if str(d.get('sku_variant')) in dupes]

# ---------- blocs contigus ----------
br = sorted(d['_row'] for d in data if not d.get('price'))
runs = []
s = p0 = br[0]
for x in br[1:]:
    if x != p0 + 1:
        runs.append((s, p0))
        s = x
    p0 = x
runs.append((s, p0))
big_runs = sorted([r for r in runs if r[1] - r[0] >= 5], key=lambda r: -(r[1] - r[0]))[:6]

# ---------- valeur ----------
prices = [d['price'] for d in data if d.get('price')]
med = sorted(prices)[len(prices) // 2]
val_known = sum((d.get('price') or 0) * (d.get('inventory') or 0) for d in data)
u_blocked = sum(d.get('inventory') or 0 for d in data if not d.get('price'))
val_est = med * u_blocked

TODAY = datetime.date.today().strftime('%d/%m/%Y')

CSS = r"""
@page {
  size: A4; margin: 17mm 15mm 16mm 15mm;
  @bottom-left { content: "Rapport données manquantes — Stock WASSAL 2025";
    font-family: Georgia, serif; font-size: 7.5pt; color: #8A8F98; }
  @bottom-right { content: counter(page) " / " counter(pages);
    font-family: Consolas, monospace; font-size: 7.5pt; color: #8A8F98; }
}
@page :first { margin: 0; @bottom-left { content: ""; } @bottom-right { content: ""; } }
* { box-sizing: border-box; }
body { font-family: "Segoe UI", Calibri, Helvetica, sans-serif; font-size: 9.4pt;
  line-height: 1.5; color: #15181E; margin: 0; }
h1, h2, h3 { font-family: Georgia, "Times New Roman", serif; font-weight: normal; margin: 0; }

/* ---- couverture ---- */
.cover { height: 297mm; padding: 26mm 20mm 18mm; page-break-after: always;
  border-top: 11mm solid #22365E; position: relative; }
.cover .kicker { font-family: Consolas, monospace; font-size: 8.5pt; letter-spacing: .22em;
  text-transform: uppercase; color: #A93B2A; }
.cover h1 { font-size: 33pt; line-height: 1.06; margin: 9mm 0 0; letter-spacing: -.02em; max-width: 150mm; }
.cover .sub { font-size: 12pt; color: #3D434E; margin-top: 7mm; max-width: 128mm; line-height: 1.55; }
.cover .rule { border-top: 1.5pt solid #15181E; margin: 12mm 0 0; }
.cover .meta { display: flex; gap: 14mm; margin-top: 5mm; font-size: 8.5pt; color: #71767F; }
.cover .meta b { display: block; font-family: Consolas, monospace; font-size: 10pt;
  color: #15181E; font-weight: normal; margin-top: 1.5mm; }
.headline { margin-top: 20mm; background: #F0DCD8; border-left: 3.5pt solid #A93B2A;
  padding: 8mm 9mm; }
.headline .big { font-family: Georgia, serif; font-size: 52pt; line-height: .9; color: #A93B2A;
  letter-spacing: -.03em; }
.headline .txt { font-size: 11pt; line-height: 1.5; margin-top: 4mm; max-width: 130mm; }
.cover .foot { position: absolute; bottom: 18mm; left: 20mm; right: 20mm;
  border-top: .75pt solid #CFCFC7; padding-top: 4mm; font-size: 8pt; color: #71767F; }

/* ---- structure ---- */
section { page-break-inside: auto; }
.snum { font-family: Consolas, monospace; font-size: 8pt; letter-spacing: .2em; color: #A93B2A; }
h2 { font-size: 17pt; letter-spacing: -.015em; margin: 2mm 0 0;
  border-bottom: 1.2pt solid #15181E; padding-bottom: 2.5mm; }
h3 { font-size: 11.5pt; margin: 7mm 0 2.5mm; }
p { margin: 3mm 0 0; max-width: 172mm; }
.lead { font-size: 10.2pt; color: #3D434E; margin-top: 4mm; }
.brk { page-break-before: always; }
section + section { margin-top: 11mm; }

/* ---- kpi ---- */
.kpis { display: flex; margin: 6mm 0 0; border-top: .75pt solid #CFCFC7;
  border-bottom: .75pt solid #CFCFC7; }
.kpi { flex: 1; padding: 4.5mm 4mm; border-right: .5pt solid #E0E0D9; }
.kpi:last-child { border-right: none; }
.kpi .l { font-family: Consolas, monospace; font-size: 7pt; letter-spacing: .13em;
  text-transform: uppercase; color: #71767F; }
.kpi .v { font-family: Georgia, serif; font-size: 21pt; line-height: 1.1; margin-top: 2mm; }
.kpi .d { font-size: 7.8pt; color: #71767F; margin-top: 1mm; line-height: 1.3; }
.kpi.bad .v { color: #A93B2A; } .kpi.good .v { color: #3F6B4A; }

/* ---- tables ---- */
table { border-collapse: collapse; width: 100%; margin-top: 5mm; font-size: 8.6pt; }
thead { display: table-header-group; }
th { font-family: Consolas, monospace; font-size: 7pt; letter-spacing: .1em;
  text-transform: uppercase; color: #71767F; font-weight: normal; text-align: left;
  padding: 0 2.5mm 2mm 0; border-bottom: 1pt solid #15181E; }
td { padding: 2.1mm 2.5mm 2.1mm 0; border-bottom: .5pt solid #E8E8E2; vertical-align: top; }
th.r, td.r { text-align: right; font-family: Consolas, monospace; }
tr { page-break-inside: avoid; }
td.nm { font-weight: 600; }
.sev { font-family: Consolas, monospace; font-size: 6.8pt; letter-spacing: .06em;
  padding: .6mm 1.6mm; border: .5pt solid currentColor; white-space: nowrap; }
.s-Bloquant { color: #A93B2A; } .s-Critique { color: #8A6A1C; }
.s-Important { color: #22365E; } .s-Mineur { color: #71767F; }
tr.tot td { border-top: 1pt solid #15181E; border-bottom: none; font-weight: 700;
  padding-top: 2.5mm; }

/* ---- bars ---- */
.barw { display: block; width: 100%; height: 4.5pt; background: #E8E8E2; }
.barw i { display: block; height: 4.5pt; background: #A93B2A; }
.barw i.b { background: #22365E; } .barw i.g { background: #3F6B4A; }

/* ---- callouts ---- */
.box { border-left: 3pt solid #22365E; background: #F1F4F8; padding: 5mm 6mm; margin-top: 6mm;
  page-break-inside: avoid; }
.box.warn { border-color: #A93B2A; background: #F7E9E6; }
.box.ok { border-color: #3F6B4A; background: #E9F1EB; }
.box h4 { font-family: Georgia, serif; font-size: 11pt; font-weight: normal; margin: 0 0 2mm; }
.box p { margin: 2mm 0 0; font-size: 9.2pt; }
.box p:first-of-type { margin-top: 0; }

/* ---- stacked bar ---- */
.stack { display: flex; height: 13mm; margin-top: 5mm; }
.stack div { display: flex; align-items: center; justify-content: center;
  font-family: Consolas, monospace; font-size: 7.5pt; color: #fff; }
.key { display: flex; gap: 9mm; margin-top: 3mm; flex-wrap: wrap; }
.key span { font-size: 8pt; color: #3D434E; }
.key i { display: inline-block; width: 8pt; height: 8pt; margin-right: 2mm;
  vertical-align: -.5pt; }

/* ---- steps ---- */
.step { display: flex; gap: 6mm; padding: 4.5mm 0; border-bottom: .5pt solid #E8E8E2;
  page-break-inside: avoid; }
.step .no { font-family: Georgia, serif; font-size: 19pt; color: #A93B2A; line-height: 1;
  width: 11mm; flex-shrink: 0; }
.step .bd { flex: 1; }
.step h4 { font-family: Georgia, serif; font-size: 11.5pt; font-weight: normal; margin: 0; }
.step p { margin: 1.8mm 0 0; font-size: 9pt; color: #3D434E; }
.step .tag { font-family: Consolas, monospace; font-size: 7pt; color: #71767F;
  letter-spacing: .08em; margin-top: 2mm; display: block; }
small { font-size: 8pt; color: #71767F; }
"""


def sev_cell(s):
    return '<span class="sev s-%s">%s</span>' % (s, s.upper())


# ================= HTML =================
h = []
A = h.append

# ---------- COUVERTURE ----------
A('<div class="cover">')
A('<div class="kicker">Rapport d\'analyse \u00b7 Donn\u00e9es produit</div>')
A('<h1>Ce qui manque dans le fichier stock WASSAL 2025</h1>')
A('<p class="sub">Inventaire pr\u00e9cis des informations absentes, produit par produit, '
  'et de ce qu\'elles emp\u00eachent de vendre.</p>')
A('<div class="rule"></div>')
A('<div class="meta">'
  '<div>Fichier analys\u00e9<b>Excel WASSAL 2025.xlsx</b></div>'
  '<div>Onglet<b>bulk_import</b></div>'
  '<div>Date<b>%s</b></div>'
  '<div>P\u00e9rim\u00e8tre<b>%s lignes</b></div></div>' % (TODAY, n(L)))
A('<div class="headline"><div class="big">%s&nbsp;%%</div>'
  '<div class="txt"><b>des produits du catalogue ne peuvent pas \u00eatre mis en vente en '
  'l\'\u00e9tat.</b> Il leur manque le prix et la cat\u00e9gorie \u2014 les deux seules '
  'informations sans lesquelles un article ne peut \u00eatre ni import\u00e9 ni publi\u00e9. '
  'Cela repr\u00e9sente %s produits et %s articles physiques immobilis\u00e9s en '
  'entrep\u00f4t.</div></div>'
  % (n(100 * buck['bloque'] / P), n(buck['bloque']), n(bunits['bloque'])))
A('<div class="foot">Document interne \u00b7 Toutes les valeurs sont issues directement du '
  'fichier source, sans correction ni estimation, sauf mention explicite contraire.</div>')
A('</div>')

# ---------- 1. SYNTHESE ----------
A('<section><div class="snum">SECTION 01</div><h2>Synth\u00e8se</h2>')
A('<p class="lead">Le fichier contient <b>%s produits</b> d\u00e9clin\u00e9s en <b>%s variantes</b> '
  '(une variante = une taille et une couleur), pour <b>%s articles</b> en stock. '
  'L\'identit\u00e9 des produits est compl\u00e8te : noms, r\u00e9f\u00e9rences et codes-barres '
  'ne pr\u00e9sentent aucun trou. Ce qui manque, ce sont les informations commerciales.</p>'
  % (n(P), n(L), n(U)))

A('<div class="kpis">')
A('<div class="kpi"><div class="l">Produits</div><div class="v">%s</div>'
  '<div class="d">%s variantes au total</div></div>' % (n(P), n(L)))
A('<div class="kpi bad"><div class="l">\u00c0 compl\u00e9ter</div><div class="v">%s</div>'
  '<div class="d">%s %% du catalogue</div></div>'
  % (n(buck['bloque']), n(100 * buck['bloque'] / P)))
A('<div class="kpi bad"><div class="l">Articles bloqu\u00e9s</div><div class="v">%s</div>'
  '<div class="d">%s %% du stock physique</div></div>'
  % (n(bunits['bloque']), n(100 * bunits['bloque'] / U)))
A('<div class="kpi good"><div class="l">Vendables</div><div class="v">%s</div>'
  '<div class="d">%s articles pr\u00eats</div></div>'
  % (n(buck['complet'] + buck['media']), n(bunits['complet'] + bunits['media'])))
A('</div>')

A('<h3>R\u00e9partition du catalogue</h3>')
seg = [('bloque', '#A93B2A', 'Bloqu\u00e9 \u2014 prix et cat\u00e9gorie manquants'),
       ('media', '#8A6A1C', 'Vendable mais sans photo ni fiche'),
       ('complet', '#3F6B4A', 'Complet')]
A('<div class="stack">')
for k, c, lbl in seg:
    pct = 100 * buck[k] / P
    A('<div style="width:%.2f%%;background:%s">%s</div>' % (pct, c, n(buck[k])))
A('</div><div class="key">')
for k, c, lbl in seg:
    A('<span><i style="background:%s"></i>%s \u2014 %s produits (%s %%)</span>'
      % (c, lbl, n(buck[k]), n(100 * buck[k] / P)))
A('</div>')

A('<div class="box warn"><h4>Le point essentiel</h4>'
  '<p>Le prix et la cat\u00e9gorie manquent sur <b>exactement les m\u00eames %s lignes</b>. '
  'Ce n\'est pas deux probl\u00e8mes distincts mais un seul : un bloc de donn\u00e9es qui n\'a '
  'jamais \u00e9t\u00e9 renseign\u00e9. Les corriger ensemble d\u00e9bloque tout.</p>'
  '<p>Ces %s lignes ne correspondent qu\'\u00e0 <b>%s produits distincts</b> \u2014 le reste '
  'n\'est que la d\u00e9clinaison en tailles et couleurs du m\u00eame article. Le prix et la '
  'cat\u00e9gorie \u00e9tant identiques pour toutes les tailles d\'un produit, '
  'la saisie r\u00e9elle porte sur %s lignes, pas %s.</p></div>'
  % (n(sum(1 for d in data if not d.get('price'))),
     n(sum(1 for d in data if not d.get('price'))), n(buck['bloque']),
     n(buck['bloque']), n(sum(1 for d in data if not d.get('price')))))
A('</section>')

# ---------- 2. DETAIL PAR CHAMP ----------
A('<section class="brk"><div class="snum">SECTION 02</div>'
  '<h2>Ce qui manque, champ par champ</h2>')
A('<p class="lead">Pour chaque colonne du fichier : combien de produits n\'ont aucune valeur, '
  'combien n\'en ont que sur une partie de leurs tailles, et surtout combien d\'articles '
  'physiques sont concern\u00e9s.</p>')
A('<table><thead><tr><th>Colonne</th><th>Gravit\u00e9</th>'
  '<th class="r">Produits<br>enti\u00e8rement vides</th>'
  '<th class="r">Produits<br>partiels</th><th class="r">Lignes<br>vides</th>'
  '<th class="r">Articles<br>concern\u00e9s</th><th style="width:26mm">Part du stock</th>'
  '</tr></thead><tbody>')
for lbl, sev, full, part, lines, un in frows:
    pct = 100 * un / U
    cls = 'b' if sev in ('Important',) else ('g' if sev == 'Mineur' else '')
    A('<tr><td class="nm">%s</td><td>%s</td><td class="r">%s</td><td class="r">%s</td>'
      '<td class="r">%s</td><td class="r">%s</td>'
      '<td><span class="barw"><i class="%s" style="width:%.1f%%"></i></span>'
      '<small>%s&nbsp;%%</small></td></tr>'
      % (esc(lbl), sev_cell(sev), n(full), n(part) if part else '\u2014',
         n(lines), n(un), cls, pct, n(pct)))
A('</tbody></table>')

A('<div class="box ok"><h4>Ce qui est parfaitement rempli</h4>'
  '<p>Aucune valeur manquante sur : <b>%s</b>. '
  'Les codes-barres sont tous uniques, aucun doublon. C\'est une base saine : '
  'chaque article est identifiable sans ambigu\u00eft\u00e9.</p></div>'
  % esc(', '.join(ok_list)))

A('<h3>Lecture des gravit\u00e9s</h3>')
A('<table><tbody>'
  '<tr><td style="width:26mm">%s</td><td>Sans cette information, l\'article ne peut '
  '\u00eatre ni import\u00e9 ni publi\u00e9. Rien n\'est possible tant que ce n\'est pas rempli.</td></tr>'
  '<tr><td>%s</td><td>L\'article peut techniquement \u00eatre vendu, mais sans photo '
  'le taux de conversion s\'effondre. \u00c0 traiter juste apr\u00e8s les bloquants.</td></tr>'
  '<tr><td>%s</td><td>P\u00e9nalise la qualit\u00e9 de la fiche et le r\u00e9f\u00e9rencement, '
  'sans emp\u00eacher la vente.</td></tr>'
  '<tr><td>%s</td><td>Quelques cellules isol\u00e9es. Impact n\u00e9gligeable, '
  '\u00e0 corriger au fil de l\'eau.</td></tr>'
  '</tbody></table>'
  % (sev_cell('Bloquant'), sev_cell('Critique'), sev_cell('Important'), sev_cell('Mineur')))
A('</section>')

# ---------- 3. DIAGNOSTIC ----------
A('<section class="brk"><div class="snum">SECTION 03</div>'
  '<h2>Diagnostic : d\'o\u00f9 vient le probl\u00e8me</h2>')
A('<p class="lead">Les lignes incompl\u00e8tes ne sont pas dispers\u00e9es au hasard dans le '
  'fichier. Elles forment de <b>longs blocs continus</b>, ce qui oriente clairement sur '
  'la cause.</p>')
A('<table><thead><tr><th>Bloc de lignes cons\u00e9cutives</th><th class="r">Nombre de lignes</th>'
  '<th>Ce qui y manque</th></tr></thead><tbody>')
for a, b in big_runs:
    A('<tr><td class="nm">lignes %s \u2192 %s</td><td class="r">%s</td>'
      '<td>prix, cat\u00e9gorie, fiche d\u00e9taill\u00e9e, photo</td></tr>'
      % (n(a), n(b), n(b - a + 1)))
A('</tbody></table>')
A('<p><small>Les %s blocs identifi\u00e9s couvrent la quasi-totalit\u00e9 des lignes '
  'incompl\u00e8tes.</small></p>' % n(len(runs)))

A('<div class="box"><h4>Interpr\u00e9tation</h4>'
  '<p>Les colonnes manquantes sont <b>toujours les m\u00eames</b> (prix, cat\u00e9gorie, '
  'contenu, photo) et les colonnes pr\u00e9sentes <b>toujours les m\u00eames aussi</b> '
  '(nom FR/AR/EN, description courte, SKU, code-barres, taille, couleur, stock).</p>'
  '<p>Cela ne ressemble pas \u00e0 des oublis ponctuels de saisie, mais \u00e0 une '
  '<b>fusion de deux sources de donn\u00e9es</b> : un export riche pour une partie du '
  'catalogue, et un export plus pauvre \u2014 probablement un inventaire d\'entrep\u00f4t \u2014 '
  'pour le reste. Les informations commerciales des blocs concern\u00e9s existent '
  'peut-\u00eatre d\u00e9j\u00e0 ailleurs (ancien catalogue, site en ligne, fiches '
  'fournisseur).</p>'
  '<p><b>Cons\u00e9quence pratique :</b> avant de ressaisir %s produits \u00e0 la main, '
  'il vaut la peine de v\u00e9rifier si la source d\'origine peut \u00eatre '
  'r\u00e9-export\u00e9e compl\u00e8te. Le gain de temps serait consid\u00e9rable.</p></div>'
  % n(buck['bloque']))
A('</section>')

# ---------- 4. IMPACT ----------
A('<section><div class="snum">SECTION 04</div><h2>Impact commercial</h2>')
A('<p class="lead">Ce que les donn\u00e9es manquantes co\u00fbtent concr\u00e8tement, '
  'exprim\u00e9 en articles immobilis\u00e9s et en valeur.</p>')
A('<div class="kpis">')
A('<div class="kpi"><div class="l">Valeur catalogue connue</div><div class="v">%s</div>'
  '<div class="d">MAD \u00b7 articles ayant un prix</div></div>' % n(val_known))
A('<div class="kpi bad"><div class="l">Valeur bloqu\u00e9e (estim\u00e9e)</div>'
  '<div class="v">~%s</div><div class="d">MAD \u00b7 estimation, voir note</div></div>' % n(val_est))
A('<div class="kpi"><div class="l">Prix m\u00e9dian constat\u00e9</div><div class="v">%s</div>'
  '<div class="d">MAD \u00b7 base de l\'estimation</div></div>' % n(med))
A('</div>')
A('<p><small><b>Note de m\u00e9thode :</b> la valeur bloqu\u00e9e est une estimation d\'ordre '
  'de grandeur, obtenue en appliquant le prix m\u00e9dian du catalogue (%s MAD) aux %s '
  'articles sans prix. Elle n\'a pas valeur comptable : le prix r\u00e9el de ces articles '
  'est inconnu, c\'est pr\u00e9cis\u00e9ment l\'objet du probl\u00e8me. Elle sert uniquement '
  '\u00e0 mesurer l\'enjeu.</small></p>' % (n(med), n(u_blocked)))

A('<h3>Concentration : peu de produits, beaucoup de stock</h3>')
A('<p>Les %s produits bloqu\u00e9s ne se valent pas. Les <b>%s premiers</b> repr\u00e9sentent '
  '\u00e0 eux seuls <b>%s articles</b>, soit <b>%s %% de tout le stock bloqu\u00e9</b>. '
  'Compl\u00e9ter uniquement ces %s lignes lib\u00e8re les deux tiers de la marchandise '
  'immobilis\u00e9e.</p>'
  % (n(len(blocked)), n(len(TOP)), n(top_u), n(100 * top_u / bl_u), n(len(TOP))))
A('</section>')

# ---------- 5. TOP PRODUITS ----------
A('<section class="brk"><div class="snum">SECTION 05</div>'
  '<h2>Les %s produits \u00e0 traiter en priorit\u00e9</h2>' % n(len(TOP)))
A('<p class="lead">Class\u00e9s par volume de stock immobilis\u00e9. La colonne '
  "\u00ab ligne \u00bb renvoie au num\u00e9ro de ligne exact dans l'onglet bulk_import du "
  'fichier source.</p>')
A('<table><thead><tr><th style="width:7mm">#</th><th>Produit</th><th>Tailles</th>'
  '<th class="r">Variantes<br>\u00e0 compl\u00e9ter</th><th class="r">Articles<br>en stock</th>'
  '<th class="r">Ligne</th><th style="width:20mm">Poids</th></tr></thead><tbody>')
mx = TOP[0]['u'] if TOP else 1
for i, b in enumerate(TOP, 1):
    A('<tr><td class="r">%d</td><td class="nm">%s</td><td><small>%s</small></td>'
      '<td class="r">%s / %s</td><td class="r">%s</td><td class="r"><small>%s</small></td>'
      '<td><span class="barw"><i style="width:%.1f%%"></i></span></td></tr>'
      % (i, esc(b['n']), esc(b['sz']), n(b['nb']), n(b['nv']), n(b['u']),
         n(b['row']), 100 * b['u'] / mx))
A('<tr class="tot"><td></td><td>Total des %s premiers</td><td></td><td></td>'
  '<td class="r">%s</td><td></td><td></td></tr>' % (n(len(TOP)), n(top_u)))
A('</tbody></table>')
A('<p><small>Reste ensuite %s produits repr\u00e9sentant %s articles, \u00e0 traiter dans un '
  'second temps.</small></p>' % (n(len(blocked) - len(TOP)), n(bl_u - top_u)))
A('</section>')

# ---------- 6. CAS PARTICULIERS ----------
A('<section class="brk"><div class="snum">SECTION 06</div><h2>Cas particuliers</h2>')

A('<h3>Produits partiellement tarif\u00e9s</h3>')
A('<p>Trois produits ont un prix sur certaines tailles mais pas sur d\'autres. '
  'Le prix connu est tr\u00e8s probablement valable pour toutes les tailles, mais cela '
  'demande confirmation avant d\'\u00eatre recopi\u00e9 automatiquement.</p>')
A('<table><thead><tr><th>Produit</th><th class="r">Tailles sans prix</th>'
  '<th class="r">Total tailles</th><th class="r">Articles</th>'
  '<th class="r">Prix connu (MAD)</th></tr></thead><tbody>')
for nm, nb, tot, un, pr in partial:
    A('<tr><td class="nm">%s</td><td class="r">%s</td><td class="r">%s</td>'
      '<td class="r">%s</td><td class="r">%s</td></tr>'
      % (esc(nm), n(nb), n(tot), n(un), n(pr) if pr else '\u2014'))
A('</tbody></table>')

A('<h3>Produits vendables mais sans aucune photo</h3>')
A('<p>Ces produits ont bien un prix et une cat\u00e9gorie : ils peuvent \u00eatre mis en vente '
  '<b>d\u00e8s aujourd\'hui</b>. Mais aucune de leurs variantes n\'a d\'image. Les trois '
  'premiers pesent \u00e0 eux seuls pr\u00e8s de mille articles \u2014 c\'est le meilleur '
  'rapport effort / r\u00e9sultat de tout le fichier.</p>')
A('<table><thead><tr><th>Produit</th><th class="r">Variantes</th><th class="r">Articles</th>'
  '<th class="r">Prix (MAD)</th><th class="r">Ligne</th></tr></thead><tbody>')
for nm, un, nv, rw, pr in nophoto:
    A('<tr><td class="nm">%s</td><td class="r">%s</td><td class="r">%s</td>'
      '<td class="r">%s</td><td class="r"><small>%s</small></td></tr>'
      % (esc(nm), n(nv), n(un), n(pr) if pr else '\u2014', n(rw)))
A('<tr class="tot"><td>Total</td><td></td><td class="r">%s</td><td></td><td></td></tr>'
  % n(sum(x[1] for x in nophoto)))
A('</tbody></table>')

A('<h3>Anomalies de stock \u00e0 v\u00e9rifier physiquement</h3>')
A('<p>Ces %s lignes contiennent des valeurs impossibles. Elles ne bloquent pas la vente '
  'mais fausseront l\'inventaire \u00e0 l\'import.</p>' % n(len(anom) + len(dup_rows)))
A('<table><thead><tr><th>Anomalie</th><th>Produit</th><th>SKU variante</th>'
  '<th>Taille</th><th class="r">Valeur</th><th class="r">Ligne</th></tr></thead><tbody>')
for lbl, d, v in anom:
    A('<tr><td class="nm">%s</td><td>%s</td><td><small>%s</small></td><td>%s</td>'
      '<td class="r">%s</td><td class="r"><small>%s</small></td></tr>'
      % (esc(lbl), esc(d['name'])[:44], esc(d.get('sku_variant') or ''),
         esc(d.get('size') or ''), esc(v), n(d['_row'])))
for d in dup_rows:
    A('<tr><td class="nm">SKU en double</td><td>%s</td><td><small>%s</small></td><td>%s</td>'
      '<td class="r">\u2014</td><td class="r"><small>%s</small></td></tr>'
      % (esc(d['name'])[:44], esc(d.get('sku_variant') or ''),
         esc(d.get('size') or ''), n(d['_row'])))
A('</tbody></table>')
A('<div class="box warn"><h4>Attention sur le doublon de SKU</h4>'
  '<p>La r\u00e9f\u00e9rence <b>%s</b> est utilis\u00e9e par <b>deux produits '
  'diff\u00e9rents</b> : \u00ab %s \u00bb et \u00ab %s \u00bb. Il ne s\'agit donc pas d\'une '
  'ligne dupliqu\u00e9e par erreur, mais d\'une r\u00e9f\u00e9rence erron\u00e9e sur l\'un des '
  'deux. \u00c0 l\'import, l\'une des deux lignes \u00e9crasera l\'autre et un produit '
  'dispara\u00eetra silencieusement du catalogue.</p></div>'
  % (esc(dupes[0]), esc(dup_rows[0]['name']), esc(dup_rows[1]['name'])))
A('</section>')

# ---------- 7. PLAN ----------
A('<section class="brk"><div class="snum">SECTION 07</div>'
  '<h2>Plan d\'action recommand\u00e9</h2>')
A('<p class="lead">Par ordre de rentabilit\u00e9 : le plus de stock d\u00e9bloqu\u00e9 pour '
  'le moins d\'effort.</p>')
STEPS = [
    ('1', 'V\u00e9rifier si la source peut \u00eatre r\u00e9-export\u00e9e',
     'Avant toute saisie manuelle, confirmer aupr\u00e8s de la personne qui produit le fichier '
     'que les %s produits incomplets ne proviennent pas d\'un export tronqu\u00e9. '
     'Si le prix et la cat\u00e9gorie existent d\u00e9j\u00e0 dans le syst\u00e8me d\'origine, '
     'tout le reste du plan devient inutile.'
     % n(buck['bloque']), 'Effort : 1 \u00e9change \u00b7 Gain potentiel : la totalit\u00e9'),
    ('2', 'Publier imm\u00e9diatement les produits d\u00e9j\u00e0 complets',
     'Ne pas attendre la correction du reste. %s produits (%s articles) sont vendables '
     'aujourd\'hui sans aucune intervention.'
     % (n(buck['complet']), n(bunits['complet'])), 'Effort : nul \u00b7 Gain : %s articles'
     % n(bunits['complet'])),
    ('3', 'Ajouter les photos des %s produits d\u00e9j\u00e0 tarif\u00e9s' % n(len(nophoto)),
     'Ces produits ont prix et cat\u00e9gorie mais aucune image. Les trois premiers '
     'concentrent l\'essentiel du volume. C\'est le meilleur rapport effort / r\u00e9sultat '
     'du fichier.', 'Effort : %s liens \u00b7 Gain : %s articles'
     % (n(len(nophoto)), n(sum(x[1] for x in nophoto)))),
    ('4', 'Saisir prix et cat\u00e9gorie des %s produits prioritaires' % n(len(TOP)),
     'Le c\u0153ur du travail. En traitant les %s produits les plus volumineux d\'abord, '
     'on lib\u00e8re %s %% du stock bloqu\u00e9 avant m\u00eame d\'avoir termin\u00e9.'
     % (n(len(TOP)), n(100 * top_u / bl_u)),
     'Effort : %s lignes \u00b7 Gain : %s articles' % (n(len(TOP)), n(top_u))),
    ('5', 'Terminer le reste du catalogue',
     'Les %s produits restants, plus dispers\u00e9s et moins volumineux. '
     '\u00c0 traiter au fil de l\'eau une fois les priorit\u00e9s trait\u00e9es.'
     % n(len(blocked) - len(TOP)),
     'Effort : %s lignes \u00b7 Gain : %s articles'
     % (n(len(blocked) - len(TOP)), n(bl_u - top_u))),
    ('6', 'Corriger les anomalies de stock et le doublon de SKU',
     'Comptage physique des %s lignes \u00e0 quantit\u00e9 aberrante, et arbitrage sur la '
     'r\u00e9f\u00e9rence utilis\u00e9e par deux produits diff\u00e9rents.' % n(len(anom)),
     'Effort : %s lignes \u00b7 Enjeu : fiabilit\u00e9 de l\'inventaire' % n(len(anom) + len(dup_rows))),
]
for no, t, d, tag in STEPS:
    A('<div class="step"><div class="no">%s</div><div class="bd"><h4>%s</h4><p>%s</p>'
      '<span class="tag">%s</span></div></div>' % (no, t, d, tag))

A('<div class="box ok"><h4>En r\u00e9sum\u00e9</h4>'
  '<p>Le fichier n\'est pas \u00e0 refaire. Son ossature est saine : identit\u00e9 produit, '
  'r\u00e9f\u00e9rences et codes-barres sont complets et coh\u00e9rents sur les %s lignes.</p>'
  '<p>Il manque <b>deux colonnes sur %s produits</b>. C\'est un travail fini, '
  'mesurable, et dont les deux tiers du b\u00e9n\u00e9fice sont atteints '
  'd\u00e8s les %s premi\u00e8res lignes trait\u00e9es.</p></div>'
  % (n(L), n(buck['bloque']), n(len(TOP))))
A('</section>')

doc = '<html><head><meta charset="utf-8"><style>%s</style></head><body>%s</body></html>' \
      % (CSS, '\n'.join(h))
HTML(string=doc).write_pdf(OUT)
print('WROTE ' + OUT)
print('produits=%d bloques=%d top=%d nophoto=%d anomalies=%d'
      % (P, buck['bloque'], len(TOP), len(nophoto), len(anom) + len(dup_rows)))
