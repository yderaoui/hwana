# -*- coding: utf-8 -*-
"""Version lisible du fichier d'origine. Aucune donnee modifiee, seulement reorganisee."""
import openpyxl, collections
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = 'c:/Users/Asus/Desktop/alsamah/Excel WASSAL 2025.xlsx'
OUT = 'c:/Users/Asus/Desktop/alsamah/WASSAL_2025_LISIBLE.xlsx'

wb = openpyxl.load_workbook(SRC, data_only=True)
rows = list(wb['bulk_import'].iter_rows(values_only=True))
hdr = [h for h in rows[0] if h]
data = []
for i, r in enumerate(rows[1:], start=2):
    if any(x is not None for x in r[:20]):
        d = dict(zip(hdr, r))
        d['_row'] = i
        data.append(d)

cats = {}
for r in wb['category ids'].iter_rows(min_row=4, values_only=True):
    if r[1] is not None:
        cats[str(r[1]).strip()] = str(r[2]).strip()
brands = {}
for r in wb['brand ids'].iter_rows(min_row=4, values_only=True):
    if r[2] is not None:
        brands[str(r[2]).strip()] = str(r[3]).strip()


def catname(d):
    v = d.get('categ_id')
    return cats.get(str(v).strip(), '') if v is not None else ''


def brandname(d):
    v = d.get('product_brand_id')
    return brands.get(str(v).strip(), '') if v is not None else ''


# regroupement par produit, ordre d'apparition conserve
by = collections.OrderedDict()
for d in data:
    by.setdefault(d['name'], []).append(d)

INK = '1F2937'
F_HDR = PatternFill('solid', fgColor=INK)
FONT_HDR = Font(color='FFFFFF', bold=True, size=10)
BAND_A = PatternFill('solid', fgColor='FFFFFF')
BAND_B = PatternFill('solid', fgColor='F1F4F8')
F_EMPTY = PatternFill('solid', fgColor='FBE9E7')
thin = Side(style='thin', color='DDDDDD')
med = Side(style='thin', color='9AA3AF')
out = openpyxl.Workbook()
out.remove(out.active)


def head(ws, headers, widths, height=32):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.fill = F_HDR
        cell.font = FONT_HDR
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='center')
    ws.row_dimensions[1].height = height
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =====================================================================
# 0. SOMMAIRE
# =====================================================================
ws = out.create_sheet('Sommaire')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 78


def p(a='', b='', bold=False, size=11, color='1F2937', fill=None):
    ws.append(['', a, b])
    r = ws.max_row
    ws.cell(r, 2).font = Font(bold=True if a else bold, size=size, color=color)
    ws.cell(r, 3).font = Font(size=size - 1, color='4B5563')
    ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical='top')
    if fill:
        for k in (2, 3):
            ws.cell(r, k).fill = PatternFill('solid', fgColor=fill)
    return r


p()
r = p('Excel WASSAL 2025 — version lisible')
ws.cell(r, 2).font = Font(bold=True, size=18)
p('', 'Exactement les mêmes données que le fichier d\'origine, simplement réorganisées. '
      'Aucune valeur n\'a été modifiée, ajoutée ni supprimée.')
p()
p('Les 3 vues', '', fill='E5E7EB')
p('1. Vue produit', '287 lignes — une ligne par produit. Le résumé : catégorie, prix, stock total, '
                    'tailles et couleurs disponibles. Commencez ici.')
p('2. Variantes', '958 lignes — le détail complet taille par taille : SKU, code-barres, prix, stock, '
                  'photo. Chaque produit est séparé visuellement du suivant.')
p('3. Textes', '958 lignes — les descriptions et fiches produit en français, arabe et anglais. '
               'Sorties à part car ce sont elles qui rendaient le fichier illisible.')
p()
p('Ce qui a été ajouté', '', fill='E5E7EB')
p('Nom de catégorie', 'Le fichier d\'origine ne contient que l\'identifiant (ex. 302). '
                      'Le nom lisible correspondant a été ajouté à côté, depuis l\'onglet « category ids ».')
p('Nom de marque', 'Idem pour l\'identifiant de marque, depuis l\'onglet « brand ids ».')
p('Colonne « Ligne d\'origine »', 'Le numéro de ligne exact dans l\'onglet bulk_import du fichier de départ, '
                                  'pour pouvoir faire l\'aller-retour.')
p()
p('Repères visuels', '', fill='E5E7EB')
r = p('Fond rosé', 'Cellule vide dans le fichier d\'origine. La donnée manque réellement, '
                   'ce n\'est pas un oubli de mise en page.')
ws.cell(r, 2).fill = F_EMPTY
p('Bandes alternées', 'Chaque produit alterne blanc / bleu clair, pour voir d\'un coup d\'œil '
                      'où commence et où finit un produit.')
p('Ligne figée', 'Les en-têtes et le nom du produit restent visibles quand vous faites défiler.')
p('Filtres', 'Chaque colonne a un filtre : cliquez sur la flèche de l\'en-tête pour trier ou filtrer.')
p()
p('Le fichier en chiffres', '', fill='E5E7EB')
p('Produits', '%d' % len(by))
p('Variantes (taille × couleur)', '%d' % len(data))
p('Articles en stock', '%s' % format(int(sum(d.get('inventory') or 0 for d in data)), ',d').replace(',', ' '))
p('Marque', ', '.join(sorted(set(brandname(d) for d in data if brandname(d)))))

# =====================================================================
# 1. VUE PRODUIT
# =====================================================================
ws = out.create_sheet('1. Vue produit')
H = ['#', 'Produit', 'Nom arabe', 'Nom anglais', 'Catégorie', 'Marque', 'Prix', 'Stock total',
     'Nb variantes', 'Tailles', 'Couleurs', 'Photo', 'Ligne d\'origine']
W = [5, 40, 32, 34, 40, 14, 12, 10, 11, 26, 34, 9, 12]
head(ws, H, W)


def uniq(g, f):
    vals = []
    for x in g:
        v = x.get(f)
        if v is not None and str(v).strip() != '' and v not in vals:
            vals.append(v)
    return vals


r = 2
for i, (name, g) in enumerate(by.items(), start=1):
    band = BAND_A if i % 2 else BAND_B
    prices = uniq(g, 'price')
    prix = prices[0] if len(prices) == 1 else (' / '.join(str(x) for x in prices) if prices else None)
    cs = uniq(g, 'categ_id')
    cat = cats.get(str(cs[0]).strip(), '') if len(cs) == 1 else (
        ' / '.join(cats.get(str(x).strip(), str(x)) for x in cs) if cs else None)
    sizes = [str(x) for x in uniq(g, 'size')]
    colors = [str(x) for x in uniq(g, 'color')]
    nphoto = sum(1 for x in g if x.get('photo_urls'))
    ws.append([i, name, (uniq(g, 'name_ar') or [None])[0], (uniq(g, 'name_en') or [None])[0],
               cat, (uniq(g, 'product_brand_id') and brandname(g[0])) or None,
               prix, sum(x.get('inventory') or 0 for x in g), len(g),
               ', '.join(sizes), ', '.join(colors),
               '%d/%d' % (nphoto, len(g)), min(x['_row'] for x in g)])
    for c in range(1, len(H) + 1):
        cell = ws.cell(r, c)
        cell.fill = band
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(vertical='center', wrap_text=(c in (2, 3, 4, 5, 10, 11)))
        if cell.value is None:
            cell.fill = F_EMPTY
    ws.cell(r, 2).font = Font(bold=True, size=10)
    if str(ws.cell(r, 3).value or '').strip():
        ws.cell(r, 3).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    for c in (1, 8, 9, 12, 13):
        ws.cell(r, c).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(r, 7).number_format = '#,##0'
    ws.cell(r, 8).number_format = '#,##0'
    if nphoto == 0:
        ws.cell(r, 12).fill = F_EMPTY
        ws.cell(r, 12).font = Font(color='A93B2A')
    r += 1
ws.freeze_panes = 'C2'
ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(H)), r - 1)

# =====================================================================
# 2. VARIANTES
# =====================================================================
ws = out.create_sheet('2. Variantes')
H = ['Produit', 'Taille', 'Couleur', 'Prix', 'Prix soldé', 'Stock', 'SKU', 'SKU variante',
     'Code-barres', 'Catégorie', 'ID catég.', 'Marque', 'ID marque', 'Photo (lien)', 'Ligne d\'origine']
W = [40, 10, 18, 10, 11, 9, 18, 22, 16, 40, 10, 13, 10, 46, 12]
head(ws, H, W)
r = 2
for i, (name, g) in enumerate(by.items(), start=1):
    band = BAND_A if i % 2 else BAND_B
    first = r
    for d in g:
        ws.append([d.get('name'), d.get('size'), d.get('color'), d.get('price'), d.get('price_sale'),
                   d.get('inventory'), d.get('sku'), d.get('sku_variant'), d.get('barcode'),
                   catname(d) or None, d.get('categ_id'), brandname(d) or None,
                   d.get('product_brand_id'), d.get('photo_urls'), d['_row']])
        for c in range(1, len(H) + 1):
            cell = ws.cell(r, c)
            cell.fill = band
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical='center', wrap_text=(c in (1, 10)))
            if cell.value is None:
                cell.fill = F_EMPTY
        for c in (2, 4, 5, 6, 11, 13, 15):
            ws.cell(r, c).alignment = Alignment(horizontal='center', vertical='center')
        for c in (4, 5, 6):
            ws.cell(r, c).number_format = '#,##0'
        for c in (7, 8, 9):
            ws.cell(r, c).font = Font(name='Consolas', size=9)
        ws.cell(r, 14).font = Font(size=8, color='6B7280')
        # nom du produit affiche une seule fois par groupe
        if r > first:
            ws.cell(r, 1).value = None
            ws.cell(r, 1).fill = band
        else:
            ws.cell(r, 1).font = Font(bold=True, size=10)
        r += 1
    # trait de separation entre produits
    for c in range(1, len(H) + 1):
        cur = ws.cell(r - 1, c).border
        ws.cell(r - 1, c).border = Border(left=cur.left, right=cur.right, top=cur.top, bottom=med)
ws.freeze_panes = 'D2'
ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(H)), r - 1)

# =====================================================================
# 3. TEXTES
# =====================================================================
ws = out.create_sheet('3. Textes')
H = ['Produit', 'Taille', 'SKU variante', 'Description courte (FR)', 'Description courte (AR)',
     'Description courte (EN)', 'Fiche détaillée (FR)', 'Fiche détaillée (AR)',
     'Fiche détaillée (EN)', 'Nom arabe', 'Nom anglais', 'Ligne d\'origine']
W = [32, 8, 20, 46, 40, 46, 52, 44, 52, 30, 32, 12]
head(ws, H, W)
r = 2
for i, (name, g) in enumerate(by.items(), start=1):
    band = BAND_A if i % 2 else BAND_B
    first = r
    for d in g:
        ws.append([d.get('name'), d.get('size'), d.get('sku_variant'),
                   d.get('description_sale'), d.get('description_sale_ar'), d.get('description_sale_en'),
                   d.get('content'), d.get('content_ar'), d.get('content_en'),
                   d.get('name_ar'), d.get('name_en'), d['_row']])
        for c in range(1, len(H) + 1):
            cell = ws.cell(r, c)
            cell.fill = band
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if cell.value is None:
                cell.fill = F_EMPTY
        for c in (5, 8, 10):
            ws.cell(r, c).alignment = Alignment(vertical='top', wrap_text=True, horizontal='right')
        for c in (2, 12):
            ws.cell(r, c).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(r, 3).font = Font(name='Consolas', size=9)
        ws.row_dimensions[r].height = 58
        if r > first:
            ws.cell(r, 1).value = None
            ws.cell(r, 1).fill = band
        else:
            ws.cell(r, 1).font = Font(bold=True, size=10)
        r += 1
    for c in range(1, len(H) + 1):
        cur = ws.cell(r - 1, c).border
        ws.cell(r - 1, c).border = Border(left=cur.left, right=cur.right, top=cur.top, bottom=med)
ws.freeze_panes = 'D2'
ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(H)), r - 1)

out.save(OUT)
print('WROTE ' + OUT)
print('produits=%d variantes=%d' % (len(by), len(data)))
