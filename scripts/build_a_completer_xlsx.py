# -*- coding: utf-8 -*-
"""Genere un classeur de saisie en francais destine a la personne qui produit les fichiers."""
import openpyxl, collections
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

SRC = 'c:/Users/Asus/Desktop/alsamah/Excel WASSAL 2025.xlsx'
OUT = 'c:/Users/Asus/Desktop/alsamah/A_COMPLETER_WASSAL_2025.xlsx'

wb = openpyxl.load_workbook(SRC, data_only=True)
rows = list(wb['bulk_import'].iter_rows(values_only=True))
hdr = [h for h in rows[0] if h]
data = []
for i, r in enumerate(rows[1:], start=2):
    if any(x is not None for x in r[:20]):
        d = dict(zip(hdr, r))
        d['_row'] = i
        data.append(d)

cats = []
for r in wb['category ids'].iter_rows(min_row=4, values_only=True):
    if r[1] is not None:
        cats.append((r[1], str(r[2]).strip()))

byname = collections.OrderedDict()
for d in data:
    byname.setdefault(d['name'], []).append(d)

# ---------- styles ----------
INK = '1F2937'
F_HDR = PatternFill('solid', fgColor=INK)
FONT_HDR = Font(color='FFFFFF', bold=True, size=10)
F_LOCK = PatternFill('solid', fgColor='EFEFEF')      # ne pas toucher
F_FILL = PatternFill('solid', fgColor='FFF7CC')      # a remplir
F_AUTO = PatternFill('solid', fgColor='E8F0FE')      # calcule
F_URG = PatternFill('solid', fgColor='FDE2DD')
thin = Side(style='thin', color='D0D0D0')
BD = Border(left=thin, right=thin, top=thin, bottom=thin)

out = openpyxl.Workbook()
out.remove(out.active)


def header(ws, headers, widths, note_row=None):
    if note_row:
        ws.append(note_row)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(1, c)
            cell.font = Font(bold=True, size=9, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill('solid', fgColor='6B7280')
        ws.row_dimensions[1].height = 18
    r = ws.max_row + 1 if note_row else 1
    for c, htxt in enumerate(headers, start=1):
        cell = ws.cell(r, c, htxt)
        cell.fill = F_HDR
        cell.font = FONT_HDR
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='center')
    ws.row_dimensions[r].height = 34
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return r


# ==================================================================
# 0. LIRE D'ABORD
# ==================================================================
ws = out.create_sheet('LIRE D ABORD')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 100
ws.sheet_view.showGridLines = False


def p(txt='', bold=False, size=11, color='1F2937', fill=None, top=0):
    ws.append(['', txt])
    r = ws.max_row
    c = ws.cell(r, 2)
    c.font = Font(bold=bold, size=size, color=color)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    if fill:
        ws.cell(r, 2).fill = PatternFill('solid', fgColor=fill)
    ws.row_dimensions[r].height = None if not txt else max(16, 15 * (len(txt) // 95 + 1))
    return r


p()
p('Fichier WASSAL 2025 — ce qu\'il reste à compléter', bold=True, size=18)
p('Merci pour le fichier ! Il est presque complet. Voici précisément ce qui manque pour qu\'on puisse '
  'mettre le stock en vente.', size=11, color='4B5563')
p()
p('  LE PROBLÈME EN UNE PHRASE  ', bold=True, size=12, fill='FDE2DD')
p('672 lignes du fichier (sur 958) ont bien le nom, le SKU, le code-barres, la taille et la couleur — '
  'mais il leur manque le PRIX et la CATÉGORIE. Sans ces deux informations, le produit ne peut pas être '
  'importé ni mis en vente. Cela représente 4 794 articles physiques, soit 39 % du stock, qui dorment '
  'dans l\'entrepôt.', size=11)
p()
p('  LA BONNE NOUVELLE  ', bold=True, size=12, fill='DCFCE7')
p('Ces 672 lignes ne correspondent qu\'à 201 produits différents (les autres lignes sont juste les '
  'déclinaisons taille/couleur du même produit). Vous n\'avez donc à remplir que 201 lignes, pas 672. '
  'Le prix et la catégorie sont les mêmes pour toutes les tailles d\'un produit — je m\'occupe de les '
  'recopier sur chaque variante ensuite.', size=11)
p()
p('  CE QU\'IL FAUT FAIRE  ', bold=True, size=12, fill='E5E7EB')
p('1.  Onglet « 1. Prix et catégories » — le travail principal. 201 lignes.', bold=True)
p('     Pour chaque produit, remplissez les cases JAUNES : le prix de vente, et la catégorie choisie '
  'dans le menu déroulant. L\'identifiant de catégorie se remplit tout seul dans la colonne bleue.', color='4B5563')
p('2.  Onglet « 2. Photos et fiches » — 204 produits sans aucune photo.', bold=True)
p('     Collez le lien de la photo. Si le produit existe déjà sur alsamah.com, le lien de la page '
  'suffit. La fiche produit détaillée (matière, composition) est utile mais pas bloquante.', color='4B5563')
p('3.  Onglet « 3. Corrections stock » — 6 lignes seulement, mais importantes.', bold=True)
p('     Des quantités négatives et un SKU en double. À vérifier physiquement.', color='4B5563')
p()
p('  CODE COULEUR DES COLONNES  ', bold=True, size=12, fill='E5E7EB')
r = p('     Gris  =  déjà rempli, ne pas modifier (sert à retrouver le produit)')
ws.cell(r, 2).fill = F_LOCK
r = p('     Jaune  =  à remplir par vos soins')
ws.cell(r, 2).fill = F_FILL
r = p('     Bleu  =  calculé automatiquement, ne rien saisir')
ws.cell(r, 2).fill = F_AUTO
p()
p('  PRIORITÉ SI VOUS MANQUEZ DE TEMPS  ', bold=True, size=12, fill='E5E7EB')
p('Faites le PRIX et la CATÉGORIE d\'abord (onglet 1) : ce sont les deux seules colonnes réellement '
  'bloquantes. Les photos et les fiches produit peuvent suivre dans un second temps — on peut déjà '
  'vendre sans elles, même si ça vend moins bien.', size=11)
p()
p('Les colonnes sont triées par nombre d\'articles en stock : les premières lignes de l\'onglet 1 sont '
  'celles qui débloquent le plus de marchandise. Si vous n\'en faites que 20, faites les 20 premières.',
  size=11, color='4B5563')
p()
p('Une question sur une ligne précise ? La colonne « Ligne d\'origine » renvoie au numéro de ligne '
  'exact dans l\'onglet bulk_import du fichier Excel WASSAL 2025.xlsx.', size=10, color='6B7280')

# ==================================================================
# LISTE CATEGORIES (source du menu deroulant)
# ==================================================================
wsc = out.create_sheet('Liste categories')
wsc.append(['Nom de la categorie', 'ID'])
for c in range(1, 3):
    wsc.cell(1, c).fill = F_HDR
    wsc.cell(1, c).font = FONT_HDR
for cid, cname in cats:
    wsc.append([cname, cid])
wsc.column_dimensions['A'].width = 70
wsc.column_dimensions['B'].width = 10
wsc.freeze_panes = 'A2'
NCAT = len(cats) + 1
CATREF = "'Liste categories'!$A$2:$A$%d" % NCAT
LOOKREF = "'Liste categories'!$A$2:$B$%d" % NCAT

# ==================================================================
# 1. PRIX ET CATEGORIES  (201 produits)
# ==================================================================
todo = []
for name, g in byname.items():
    need_price = all(x.get('price') is None for x in g)
    need_cat = any(not x.get('categ_id') for x in g)
    if not (need_price or need_cat):
        continue
    units = sum(x.get('inventory') or 0 for x in g)
    known_price = next((x['price'] for x in g if x.get('price') is not None), None)
    todo.append({
        'name': name, 'g': g, 'units': units, 'need_price': need_price,
        'known_price': known_price, 'nvar': len(g),
        'row0': min(x['_row'] for x in g),
        'sizes': ', '.join(sorted(set(str(x.get('size') or '') for x in g)))[:60],
        'colors': ', '.join(sorted(set(str(x.get('color') or '') for x in g)))[:60],
        'sku': g[0].get('sku') or '',
        'desc': (g[0].get('description_sale') or '')[:120],
    })
todo.sort(key=lambda x: -x['units'])

ws = out.create_sheet('1. Prix et categories')
H = ['Priorite', 'Produit', 'Description courte', 'Tailles concernees', 'Couleurs concernees',
     'Nb variantes', 'Articles en stock', 'Ligne d origine',
     'PRIX DE VENTE (MAD)', 'CATEGORIE (menu deroulant)', 'ID categorie (auto)']
W = [9, 44, 46, 30, 30, 11, 11, 12, 17, 46, 15]
NOTE = ['', '', '', 'NE PAS MODIFIER — sert a identifier le produit', '', '', '', '',
        'A REMPLIR', '', 'AUTO']
hr = header(ws, H, W, note_row=NOTE)
ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=8)
ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10)
ws.cell(1, 2, 'NE PAS MODIFIER — sert a identifier le produit')
ws.cell(1, 9, 'A REMPLIR PAR VOS SOINS')
ws.cell(1, 11, 'AUTO')
for c, f in [(2, '6B7280'), (9, 'B7791F'), (11, '3B6FB6')]:
    ws.cell(1, c).fill = PatternFill('solid', fgColor=f)
    ws.cell(1, c).font = Font(bold=True, size=9, color='FFFFFF')
    ws.cell(1, c).alignment = Alignment(horizontal='center', vertical='center')

dv = DataValidation(type='list', formula1='=%s' % CATREF, allow_blank=True, showDropDown=False)
dv.error = 'Choisissez une categorie dans la liste proposee.'
dv.errorTitle = 'Categorie inconnue'
dv.prompt = 'Cliquez sur la fleche pour choisir la categorie.'
dv.promptTitle = 'Categorie'
ws.add_data_validation(dv)

dvp = DataValidation(type='decimal', operator='greaterThan', formula1='0', allow_blank=True)
dvp.error = 'Le prix doit etre un nombre superieur a 0 (en dirhams, sans texte).'
dvp.errorTitle = 'Prix invalide'
ws.add_data_validation(dvp)

r = hr + 1
for i, t in enumerate(todo, start=1):
    prio = 'HAUTE' if t['units'] >= 40 else 'MOYENNE' if t['units'] >= 12 else 'BASSE'
    ws.append([prio, t['name'], t['desc'], t['sizes'], t['colors'], t['nvar'], t['units'],
               t['row0'], t['known_price'], '', ''])
    for c in range(1, 12):
        ws.cell(r, c).border = BD
    for c in range(1, 9):
        ws.cell(r, c).fill = F_LOCK
    ws.cell(r, 1).fill = PatternFill(
        'solid', fgColor={'HAUTE': 'FDE2DD', 'MOYENNE': 'FFF1D6', 'BASSE': 'EFEFEF'}[prio])
    ws.cell(r, 1).font = Font(bold=True, size=9)
    ws.cell(r, 1).alignment = Alignment(horizontal='center')
    ws.cell(r, 7).number_format = '#,##0'
    for c in (9, 10):
        ws.cell(r, c).fill = F_FILL
    ws.cell(r, 9).number_format = '#,##0'
    if t['known_price'] is not None:
        ws.cell(r, 9).comment = Comment(
            "Prix deja present sur une autre taille du meme produit. "
            "Confirmez ou corrigez.", "Analyse stock")
    ws.cell(r, 11).value = ('=IF($J%d="","",IFERROR(VLOOKUP($J%d,%s,2,FALSE),"introuvable"))'
                            % (r, r, LOOKREF))
    ws.cell(r, 11).fill = F_AUTO
    ws.cell(r, 11).font = Font(color='3B6FB6')
    ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical='top')
    dv.add(ws.cell(r, 10))
    dvp.add(ws.cell(r, 9))
    r += 1
ws.freeze_panes = 'C%d' % (hr + 1)
ws.auto_filter.ref = 'A%d:K%d' % (hr, r - 1)

# ==================================================================
# 2. PHOTOS ET FICHES
# ==================================================================
photo_todo = []
for name, g in byname.items():
    if all(not x.get('photo_urls') for x in g):
        photo_todo.append({
            'name': name, 'units': sum(x.get('inventory') or 0 for x in g),
            'nvar': len(g), 'sku': g[0].get('sku') or '',
            'row0': min(x['_row'] for x in g),
            'has_content': any(x.get('content') for x in g),
            'desc': (g[0].get('description_sale') or '')[:120],
        })
photo_todo.sort(key=lambda x: -x['units'])

ws = out.create_sheet('2. Photos et fiches')
H = ['Produit', 'Description courte', 'SKU', 'Nb variantes', 'Articles en stock',
     'Ligne d origine', 'Fiche detaillee deja presente ?',
     'LIEN PHOTO (URL)', 'FICHE PRODUIT DETAILLEE (matiere, composition)']
W = [42, 46, 16, 11, 12, 12, 17, 52, 60]
hr = header(ws, H, W)
r = hr + 1
for t in photo_todo:
    ws.append([t['name'], t['desc'], t['sku'], t['nvar'], t['units'], t['row0'],
               'oui' if t['has_content'] else 'non', '', ''])
    for c in range(1, 10):
        ws.cell(r, c).border = BD
    for c in range(1, 8):
        ws.cell(r, c).fill = F_LOCK
    for c in (8, 9):
        ws.cell(r, c).fill = F_FILL
    ws.cell(r, 5).number_format = '#,##0'
    ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.cell(r, 7).alignment = Alignment(horizontal='center')
    if not t['has_content']:
        ws.cell(r, 7).font = Font(color='A93B2A', bold=True)
    r += 1
ws.freeze_panes = 'B%d' % (hr + 1)
ws.auto_filter.ref = 'A%d:I%d' % (hr, r - 1)

# ==================================================================
# 3. CORRECTIONS STOCK
# ==================================================================
ws = out.create_sheet('3. Corrections stock')
H = ['Probleme', 'Ligne d origine', 'Produit', 'SKU variante', 'Taille', 'Couleur',
     'Valeur actuelle', 'QUANTITE REELLE (a verifier en entrepot)', 'Commentaire']
W = [26, 13, 42, 22, 10, 16, 14, 24, 46]
hr = header(ws, H, W)
r = hr + 1
anomalies = []
for d in data:
    v = d.get('inventory')
    if isinstance(v, (int, float)):
        if v < 0:
            anomalies.append(('Quantite negative', d, v,
                              'Une quantite ne peut pas etre negative. Comptez le stock reel.'))
        elif v != int(v):
            anomalies.append(('Quantite a virgule', d, v,
                              'Quantite decimale impossible pour un article vendu a la piece.'))
dupes = [k for k, v in collections.Counter(str(x.get('sku_variant')) for x in data).items() if v > 1]
for d in data:
    if str(d.get('sku_variant')) in dupes:
        anomalies.append(('SKU variante en double', d, d.get('sku_variant'),
                          'Ce SKU apparait 2 fois. A l import, une des deux lignes ecrasera l autre.'))
for lbl, d, val, note in anomalies:
    ws.append([lbl, d['_row'], d['name'], d.get('sku_variant') or '', str(d.get('size') or ''),
               d.get('color') or '', val, '', note])
    for c in range(1, 10):
        ws.cell(r, c).border = BD
        ws.cell(r, c).fill = F_URG
    ws.cell(r, 8).fill = F_FILL
    ws.cell(r, 9).alignment = Alignment(wrap_text=True, vertical='top')
    r += 1
ws.freeze_panes = 'A%d' % (hr + 1)

# ==================================================================
# 4. RAPPEL DES REGLES DU FICHIER
# ==================================================================
ws = out.create_sheet('4. Regles du fichier')
H = ['Colonne', 'A quoi ca sert', 'Obligatoire ?', 'Regle a respecter', 'Etat actuel du fichier']
W = [24, 42, 14, 46, 40]
hr = header(ws, H, W)
REGLES = [
    ('name', 'Nom du produit en francais', 'Obligatoire', 'Maximum 100 caracteres', 'OK — 958/958 remplies'),
    ('name_ar / name_en', 'Nom en arabe et en anglais', 'Optionnel', 'Maximum 100 caracteres', 'OK — 958/958 remplies'),
    ('description_sale', 'Description courte affichee en boutique', 'Obligatoire', 'Maximum 200 caracteres', 'OK — 955/958'),
    ('content', 'Fiche detaillee (matiere, composition)', 'Optionnel', 'Texte libre', 'MANQUE — 692 lignes vides'),
    ('price / price_sale', 'Prix de vente en dirhams', 'Obligatoire', 'Nombre uniquement, sans devise ni espace', 'MANQUE — 654 lignes vides'),
    ('categ_id', 'Identifiant de la categorie', 'Obligatoire', 'Doit exister dans l onglet « category ids »', 'MANQUE — 672 lignes vides'),
    ('sku / sku_variant', 'Reference produit et reference declinaison', 'Obligatoire', 'sku_variant doit etre unique dans tout le fichier', '1 doublon : T3308003009-S/M'),
    ('barcode', 'Code-barres EAN', 'Obligatoire', 'Doit etre unique', 'OK — aucun doublon'),
    ('product_brand_id', 'Identifiant de la marque', 'Obligatoire', 'Doit exister dans l onglet « brand ids »', 'OK — tout en ALSAMAH (149)'),
    ('color / size', 'Couleur et taille de la declinaison', 'Obligatoire', 'Une ligne = une couleur + une taille', 'OK — 7 couleurs vides'),
    ('photo_urls', 'Lien vers la photo du produit', 'Obligatoire', 'URL complete commencant par https://', 'MANQUE — 690 lignes vides'),
    ('inventory', 'Quantite en stock', 'Obligatoire', 'Nombre entier positif ou zero', '5 anomalies (voir onglet 3)'),
]
r = hr + 1
for reg in REGLES:
    ws.append(list(reg))
    for c in range(1, 6):
        ws.cell(r, c).border = BD
        ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
    ws.cell(r, 1).font = Font(bold=True, size=10)
    if reg[4].startswith('MANQUE') or reg[4][0].isdigit():
        ws.cell(r, 5).fill = F_URG
        ws.cell(r, 5).font = Font(color='A93B2A', bold=True)
    else:
        ws.cell(r, 5).font = Font(color='3F6B4A')
    ws.row_dimensions[r].height = 30
    r += 1
ws.freeze_panes = 'A%d' % (hr + 1)

out.save(OUT)
print('WROTE ' + OUT)
print('onglet 1: %d produits | onglet 2: %d produits | onglet 3: %d lignes'
      % (len(todo), len(photo_todo), len(anomalies)))
