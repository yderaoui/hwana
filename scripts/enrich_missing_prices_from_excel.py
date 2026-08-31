from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from statistics import median
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
CATALOG = ROOT / "src" / "data" / "catalog.json"
ALSAMAH_BOOK = ROOT / "alsamah 18-08-2026 (1).xlsx"

HEX = {
    "noir": "#1d1d1c",
    "blanc": "#f5f5f0",
    "blanc-casse": "#e9e2d2",
    "beige": "#cfb491",
    "gris": "#a8aaad",
    "bleu-marine": "#17233d",
    "bleu": "#36608d",
    "bleu-ciel": "#91b9d0",
    "rose": "#c9828e",
    "fushia": "#b73577",
    "rouge": "#a7473e",
    "jaune": "#d0a53a",
    "marron": "#6f4a35",
    "orange": "#c8733d",
    "violet": "#765d87",
    "vert": "#627b54",
    "kaki": "#6f7351",
    "bordeaux": "#6e2639",
    "saumon": "#d79079",
    "colore": "#8a756f",
}

COLOR_EN = {
    "noir": "Black",
    "blanc": "White",
    "blanc-casse": "Off-white",
    "beige": "Beige",
    "gris": "Gray",
    "bleu-marine": "Navy",
    "bleu": "Blue",
    "bleu-ciel": "Sky blue",
    "rose": "Pink",
    "fushia": "Fuchsia",
    "rouge": "Red",
    "jaune": "Yellow",
    "marron": "Brown",
    "orange": "Orange",
    "violet": "Purple",
    "vert": "Green",
    "kaki": "Khaki",
    "bordeaux": "Burgundy",
    "saumon": "Salmon",
    "colore": "Colored",
}

COLOR_AR = {
    "noir": "أسود",
    "blanc": "أبيض",
    "blanc-casse": "أوف وايت",
    "beige": "بيج",
    "gris": "رمادي",
    "bleu-marine": "كحلي",
    "bleu": "أزرق",
    "bleu-ciel": "أزرق سماوي",
    "rose": "وردي",
    "fushia": "فوشيا",
    "rouge": "أحمر",
    "jaune": "أصفر",
    "marron": "بني",
    "orange": "برتقالي",
    "violet": "بنفسجي",
    "vert": "أخضر",
    "kaki": "كاكي",
    "bordeaux": "عنابي",
    "saumon": "سلموني",
    "colore": "ملون",
}

STOPWORDS = {
    "a",
    "avec",
    "d",
    "de",
    "des",
    "du",
    "en",
    "et",
    "la",
    "le",
    "les",
    "pour",
    "sans",
    "taille",
    "un",
    "with",
    "and",
    "the",
}

GENERIC_KEYS = {
    "corset",
    "legging",
    "pyjama",
    "soutien gorge",
    "slip",
}


def clean(value: Any) -> str:
    text = str(value or "").replace("\ufffd", "").strip()
    return re.sub(r"\s+", " ", text).strip(" -")


def fold(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text.casefold()).strip()


def slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "-", fold(value)).strip("-") or "produit"


def tokens(value: Any) -> list[str]:
    return [word for word in fold(value).split() if word not in STOPWORDS and not word.isdigit()]


def description_key(value: Any) -> str:
    words = tokens(value)
    return " ".join(words)


def color_key(value: Any) -> str:
    value = fold(value)
    aliases = {
        "blanche": "blanc",
        "white": "blanc",
        "noire": "noir",
        "black": "noir",
        "navy": "bleu-marine",
        "bleu marine": "bleu-marine",
        "bleu ciel": "bleu-ciel",
        "blue": "bleu",
        "pink": "rose",
        "rose fushia": "fushia",
        "fuchsia": "fushia",
        "fushia": "fushia",
        "somon": "saumon",
        "soman": "saumon",
        "blanc cass": "blanc-casse",
        "blanc casse": "blanc-casse",
        "blanc cassee": "blanc-casse",
        "off white": "blanc-casse",
        "colore": "colore",
        "coloree": "colore",
    }
    return aliases.get(value, value.replace(" ", "-") or "colore")


def number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def sale_price(regular: float | None) -> int | None:
    if regular is None:
        return None
    return int((Decimal(str(regular)) * Decimal("0.70")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def color_entry(color_id: str, image: str | None = None) -> dict[str, Any]:
    fr = color_id.replace("-", " ").capitalize()
    return {
        "id": color_id,
        "label": {"fr": fr, "ar": COLOR_AR.get(color_id, fr), "en": COLOR_EN.get(color_id, fr)},
        "hex": HEX.get(color_id, "#8a8882"),
        "image": image,
        "imageKind": "source" if image else "missing",
        "lifestyleImage": None,
    }


def category_from_name(name: str) -> str:
    n = fold(name)
    if any(word in n for word in ("garcon", "boy", "fille", "girl")):
        return "enfants"
    if any(word in n for word in ("homme", "men", "mens")):
        return "homme"
    if any(word in n for word in ("chaussette", "sock", "collant", "tight", "jambiere")):
        return "autres"
    return "femme"


def subcategory(category: str, name: str) -> str | None:
    n = fold(name)
    if category == "femme":
        if re.search(r"corset|gainant|shaper", n):
            return "corsets"
        if re.search(r"soutien|brassiere|bra", n):
            return "soutien-gorge"
        if re.search(r"body|bretelle|caraco", n):
            return "bodies"
        if re.search(r"boxer|culotte|slip", n):
            return "culottes"
        if re.search(r"lingerie|sexy|nuisette|chemise de nuit|peignoir|dentel", n):
            return "lingerie"
        if "collant" in n or "tight" in n:
            return "collants"
        if "chaussette" in n or "sock" in n:
            return "chaussettes"
        if "pyjama" in n or "homewear" in n:
            return "nuisettes"
        return "vetements"
    if category == "homme":
        if re.search(r"boxer|slip|brief", n):
            return "sous-vetements"
        if re.search(r"corset|gainant|shaper", n):
            return "corsets"
        return "hauts"
    if category == "enfants":
        if re.search(r"garcon|boy", n):
            return "garcon"
        if re.search(r"fille|girl", n):
            return "fille"
        if re.search(r"chaussette|sock", n):
            return "chaussettes"
        return "collants"
    if category == "autres":
        if re.search(r"chaussette|sock", n):
            return "chaussettes"
        if re.search(r"collant|tight|jambiere", n):
            return "collants"
    return None


def read_excel_groups() -> dict[str, list[dict[str, Any]]]:
    rows_by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
    workbook = load_workbook(ALSAMAH_BOOK, read_only=True, data_only=True)
    ws = workbook.active
    rows = ws.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    for row in rows:
        item = dict(zip(headers, row))
        desc = clean(item.get("DESCRIPTION"))
        price = number(item.get("PRIX DE VENTE"))
        qty = max(0, int(number(item.get("QUANTITE")) or 0))
        if not desc or price is None or qty <= 0:
            continue
        key = description_key(desc)
        if key:
            rows_by_key[key].append(item)
    return rows_by_key


def best_group(product: dict[str, Any], groups: dict[str, list[dict[str, Any]]]) -> tuple[str, list[dict[str, Any]]] | None:
    name = product.get("name", {}).get("fr", "")
    product_key = description_key(name)
    if not product_key or product_key in GENERIC_KEYS:
        return None
    product_tokens = set(product_key.split())
    matches: list[tuple[float, int, int, str, list[dict[str, Any]]]] = []
    for key, rows in groups.items():
        group_tokens = set(key.split())
        overlap = len(product_tokens & group_tokens)
        coverage = overlap / max(len(product_tokens), 1)
        # Product names are the customer-facing source of truth. The Excel rows
        # can include SKUs, sizes, and colors, but they must still cover almost
        # all meaningful words in the product name.
        if overlap >= 2 and coverage >= 0.75:
            matches.append((coverage, overlap, -abs(len(group_tokens) - len(product_tokens)), key, rows))
    if not matches:
        return None
    matches.sort(reverse=True)
    best = matches[0]
    same_score = [item for item in matches if item[:3] == best[:3]]
    if len({item[2] for item in same_score}) > 1:
        return None
    return best[3], best[4]


def main() -> None:
    products = json.loads(CATALOG.read_text(encoding="utf-8"))
    groups = read_excel_groups()
    promoted: list[dict[str, Any]] = []

    for product in products:
        if product.get("price") is not None and product.get("regularPrice") is not None:
            continue
        match = best_group(product, groups)
        if not match:
            continue
        key, rows = match
        prices = [number(row.get("PRIX DE VENTE")) for row in rows]
        prices = [price for price in prices if price is not None]
        regular = float(median(prices)) if prices else None
        if regular is None:
            continue

        existing_images = {
            color.get("id"): color.get("image")
            for color in product.get("colors", [])
            if color.get("id") and color.get("image")
        }
        first_image = next(iter(existing_images.values()), None)
        color_ids = list(dict.fromkeys(color_key(row.get("Color")) for row in rows))
        product["colors"] = [color_entry(color_id, existing_images.get(color_id) or first_image) for color_id in color_ids]
        variants = []
        for row in rows:
            color_id = color_key(row.get("Color"))
            size = clean(row.get("Size")) or "TU"
            barcode = clean(row.get("Code Barre")) or None
            qty = max(0, int(number(row.get("QUANTITE")) or 0))
            variants.append({
                "id": barcode or f"{product['id']}-{slug(size)}-{color_id}",
                "colorId": color_id,
                "size": size,
                "stock": qty,
                "barcode": barcode,
            })
        product["variants"] = variants
        product["variantCount"] = len(variants)
        product["stock"] = sum(variant["stock"] for variant in variants)
        product["sizes"] = sorted({variant["size"] for variant in variants})
        product["regularPrice"] = regular
        product["price"] = sale_price(regular)
        category = category_from_name(product["name"]["fr"])
        product["category"] = category
        product["categoryName"] = {
            "femme": "Femme",
            "homme": "Homme",
            "enfants": "Enfants",
            "autres": "Collants & chaussettes",
        }[category]
        product["subcategory"] = subcategory(category, product["name"]["fr"])
        product["purchasable"] = product["price"] is not None and product["stock"] > 0
        product.setdefault("missing", {})["price"] = product["regularPrice"] is None
        product.setdefault("missing", {})["category"] = False
        product.setdefault("missing", {})["image"] = not any(color.get("image") for color in product["colors"])
        product["imageStatus"] = "source" if not product["missing"]["image"] else "missing"
        product["fallbackImage"] = False
        promoted.append({"id": product["id"], "name": product["name"]["fr"], "matchedExcelKey": key, "regularPrice": regular, "price": product["price"], "stock": product["stock"]})

    CATALOG.write_text(json.dumps(products, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"promoted": len(promoted), "examples": promoted[:40]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
