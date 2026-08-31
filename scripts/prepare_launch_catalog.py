from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter, defaultdict
from decimal import Decimal, ROUND_HALF_UP
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public"
CATALOG = ROOT / "src" / "data" / "catalog.json"
ALSAMAH_BOOK = ROOT / "alsamah 18-08-2026 (1).xlsx"
ELKO_BOOK = ROOT / "Les articles disponibles ELKO depot 17-08-2026.xlsx"

HEX = {
    "noir": "#1d1d1c", "black": "#1d1d1c",
    "blanc": "#f5f5f0", "white": "#f5f5f0",
    "blanc casse": "#e9e2d2", "off-white": "#e9e2d2",
    "beige": "#cfb491", "skin": "#cfb491", "peau": "#cfb491",
    "gris": "#a8aaad", "gray": "#a8aaad", "grey": "#a8aaad", "metal": "#9fa4aa",
    "bleu marine": "#17233d", "navy": "#17233d",
    "bleu": "#36608d", "blue": "#36608d", "bleu ciel": "#91b9d0",
    "rose": "#c9828e", "pink": "#c9828e", "rose fushia": "#b73577",
    "rouge": "#a7473e", "jaune": "#d0a53a", "marron": "#6f4a35",
    "orange": "#c8733d", "violet": "#765d87", "vert": "#627b54",
    "kaki": "#6f7351", "bordeaux": "#6e2639", "saumon": "#d79079",
}

COLOR_EN = {
    "noir": "Black", "blanc": "White", "beige": "Beige", "gris": "Gray",
    "bleu marine": "Navy", "bleu": "Blue", "rose": "Pink", "rouge": "Red",
    "jaune": "Yellow", "marron": "Brown", "orange": "Orange", "violet": "Purple",
    "vert": "Green", "kaki": "Khaki", "bordeaux": "Burgundy", "saumon": "Salmon",
    "blanc casse": "Off-white", "metal": "Metal",
}

COLOR_AR = {
    "noir": "اسود", "blanc": "ابيض", "beige": "بيج", "gris": "رمادي",
    "bleu marine": "كحلي", "bleu": "ازرق", "rose": "وردي", "rouge": "احمر",
    "jaune": "اصفر", "marron": "بني", "orange": "برتقالي", "violet": "بنفسجي",
    "vert": "اخضر", "kaki": "كاكي", "bordeaux": "عنابي", "saumon": "سلموني",
    "blanc casse": "اوف وايت", "metal": "معدني",
}

DESCRIPTION_COLOR_WORDS = {
    "2", "20", "27", "beige", "black", "blanc", "blanc casse", "blanche",
    "bleu", "bleu ciel", "bleu marine", "blue", "bordou", "bordeaux", "brown",
    "colore", "fuchsia", "fushia", "green", "grey", "griis", "gris", "imprime",
    "imprimee", "ivoire", "ivory", "jaune", "lanc", "marron", "metal", "navy",
    "noir", "noire", "off white", "orange", "peau", "pink", "printed", "purple",
    "red", "rose", "rouge", "saumon", "skin", "sky blue", "soman", "somon",
    "vert", "violet", "white", "yellow",
}

DESCRIPTION_SIZE_WORDS = {
    "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14",
    "xs", "s", "m", "l", "xl", "xxl", "2xl", "3xl", "4xl", "5xl", "6xl",
    "s m", "m l", "l xl", "xl 2xl", "2xl 3xl", "3xl 4xl", "3 4xl",
    "taille libre", "free size", "tu",
}

TOKEN_ALIASES = {
    "womens": "femme", "women": "femme", "woman": "femme",
    "mens": "homme", "men": "homme", "man": "homme",
    "girls": "fille", "girl": "fille", "boys": "garcon", "boy": "garcon",
    "socks": "chaussette", "sock": "chaussette", "chaussettes": "chaussette",
    "tights": "collant", "tight": "collant", "pantyhose": "collant", "collants": "collant",
    "cotton": "coton", "sleeves": "manche", "sleeve": "manche", "manches": "manche",
    "briefs": "slip", "brief": "slip", "pregnant": "grossesse", "nursing": "allaitement",
    "seamless": "sanscouture", "dentel": "dentelle", "dantel": "dentelle",
    "printed": "imprime", "leggings": "legging", "bretelles": "bretelle",
    "shorts": "short", "hautes": "haute", "montantes": "haute",
    "longue": "long", "longues": "long", "doubled": "double",
    "boxers": "boxer", "corsets": "corset", "pyjamas": "pyjama",
}

MATCH_STOPWORDS = {
    "a", "avec", "de", "des", "du", "en", "et", "la", "le", "les", "pour",
    "sans", "taille", "the", "with", "and", "100", "96", "4",
}

PRODUCT_ALIASES = {
    "nursing-bra": "soutien-gorge-dallaitement",
    "pregnant-legging-seamless": "legging-de-grossesse-en-microfibre-sans-couture",
    "high-elasticity-legging": "legging-695067504138",
    "chaussette-en-dentelle-fille": "chaussettes-dentelle-fille",
    "chaussette-en-cotton": "chaussettes-en-coton",
    "women-diabetic-socks": "chaussettes-diabetiques-pour-femmes",
    "women-diabetics-cotton-socks": "chaussettes-diabetiques-pour-femmes",
    "doubled-boxer-corset-with-long-leg": "boxer-gainant-double-avec-jambe-longue",
    "boxer-double-corset-de-pantalon": "boxer-corset-double-a-taille-haute",
    "100-cotton-slip-low-waist-girls": "slip-fille-taille-basse",
    "pregnant-body-wide-band": "body-de-grossesse-a-large-bande",
    "girls-dantel-boxer-briefs": "boxer-fille-en-dentelle",
    "bretelles-wide-band": "body-a-bretelles-avec-bande-large",
    "pantakour": "pantakour-en-microfibre-sans-couture",
    "mens-corset-short-sleeve-undershirt": "men-s-corset-short-sleeve",
    "mens-corset-boxer": "men-s-corset-boxer",
    "bra-dentel": "bra-dentelle",
    "body-slip-dentel": "body-slip-dentelle",
    "100-cotton-no-sleeve-v-neck": "100-coton-sans-manches-col-en-v",
    "fantasia-socket-lareen": "chaussettes-fantasia-lareen",
    "leggings-fille": "legging-fille",
    "jambieres-de-leen-collant": "jambieres-leen",
    "medium-fishnet-pantyhose": "collants-fishnet",
}

REAL_COLOR_ASSET_DIRS = (
    ("generated", PUBLIC / "assets" / "catalog"),
    ("generated", PUBLIC / "assets" / "generated" / "imagegen"),
    ("generated", PUBLIC / "assets" / "generated" / "products"),
    ("source", PUBLIC / "assets" / "excel-products" / "alsamah"),
    ("source", PUBLIC / "assets" / "official"),
    ("source", PUBLIC / "assets" / "remote-products"),
    ("source", PUBLIC / "assets" / "source"),
    ("generated", PUBLIC / "assets" / "products"),
)


def clean(value: Any) -> str:
    text = str(value or "").replace("\ufffd", "").strip()
    text = re.sub(r"\s+", " ", text)
    return text.strip(" -")


def fold(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text.casefold()).strip()


def slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "-", fold(value)).strip("-") or "produit"


def color_key(value: Any) -> str:
    value = clean(value).casefold()
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"\s+", " ", value).strip()
    aliases = {
        "blanche": "blanc", "white": "blanc", "noire": "noir", "black": "noir",
        "gray": "gris", "grey": "gris", "griis": "gris", "navy": "bleu marine",
        "blue": "bleu", "pink": "rose", "soman": "saumon", "somon": "saumon",
        "lanc": "blanc", "bordou": "bordeaux", "blanc cass": "blanc casse",
        "blanc cassee": "blanc casse", "none": "non renseignee",
    }
    return aliases.get(value, value or "non renseignee")


def color_entry(name: str, image: str | None = None) -> dict[str, Any]:
    color = color_key(name)
    label_fr = "Non renseignee" if color == "non renseignee" else color.capitalize()
    return {
        "id": slug(color),
        "label": {
            "fr": label_fr,
            "ar": COLOR_AR.get(color, label_fr),
            "en": COLOR_EN.get(color, label_fr),
        },
        "hex": HEX.get(color, "#8a8882"),
        "image": image,
        "imageKind": "source" if image else "missing",
    }


def category_from_name(name: str, default: str = "femme") -> str:
    n = fold(name)
    if any(word in n for word in ("garcon", "boy", "fille", "girl", "kids", "child")):
        return "enfants"
    if any(word in n for word in ("homme", "men", "mens")):
        return "homme"
    if any(word in n for word in ("chaussette", "socks", "collant", "tight", "jambiere")):
        return "autres"
    return default


def subcategory(category: str, name: str) -> str | None:
    n = fold(name)
    if category == "femme":
        if re.search(r"corset|gainant|shaper", n):
            return "corsets"
        if re.search(r"soutien|brassiere|bra", n):
            return "soutien-gorge"
        if re.search(r"body|bretelle|caraco", n):
            return "bodies"
        if re.search(r"boxer|culotte|slip", n):
            return "culottes"
        if re.search(r"lingerie|sexy|nuisette|chemise de nuit|peignoir|dentel", n):
            return "lingerie"
        if "collant" in n or "tight" in n:
            return "collants"
        if "chaussette" in n or "sock" in n:
            return "chaussettes"
        if "pyjama" in n or "homewear" in n:
            return "nuisettes"
        return "vetements"
    if category == "homme":
        if re.search(r"boxer|slip|brief", n):
            return "sous-vetements"
        if re.search(r"corset|gainant|shaper", n):
            return "corsets"
        return "hauts"
    if category == "enfants":
        if re.search(r"garcon|boy", n):
            return "garcon"
        if re.search(r"fille|girl", n):
            return "fille"
        if re.search(r"chaussette|sock", n):
            return "chaussettes"
        return "collants"
    if category == "autres":
        if re.search(r"chaussette|sock", n):
            return "chaussettes"
        if re.search(r"collant|tight|jambiere", n):
            return "collants"
    return None


def discounted_price(regular: float | None) -> int | None:
    if regular is None:
        return None
    return int((Decimal(str(regular)) * Decimal("0.70")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def sale_price(regular: float | None, stock: float) -> int | None:
    return discounted_price(regular) if stock > 0 else None


def representative_variant(variants: list[dict[str, Any]]) -> dict[str, Any] | None:
    priced = [variant for variant in variants if variant.get("regularPrice") is not None]
    candidates = [variant for variant in priced if variant.get("stock", 0) > 0] or priced
    if not candidates:
        return None
    representative_price = Counter(variant["regularPrice"] for variant in candidates).most_common(1)[0][0]
    return next(variant for variant in candidates if variant["regularPrice"] == representative_price)


def real_color_asset(product_id: str, color_id: str) -> tuple[str, str] | None:
    stem = f"{product_id}-{color_id}"
    for image_kind, directory in REAL_COLOR_ASSET_DIRS:
        for suffix in (".png", ".jpg", ".jpeg", ".webp", ".avif"):
            path = directory / f"{stem}{suffix}"
            if path.exists():
                return "/" + path.relative_to(PUBLIC).as_posix(), image_kind
    return None


def number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def read_alsamah() -> dict[str, list[dict[str, Any]]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    if not ALSAMAH_BOOK.exists():
        return groups
    ws = load_workbook(ALSAMAH_BOOK, read_only=True, data_only=True).active
    rows = ws.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    for row in rows:
        item = dict(zip(headers, row))
        name = clean(item.get("DESCRIPTION"))
        if not name:
            continue
        groups[slug(name)].append(item)
    return groups


def normalize_alsamah_description(item: dict[str, Any]) -> str:
    description = clean(item.get("DESCRIPTION"))
    barcode = clean(item.get("Code Barre"))
    if barcode:
        for prefix in (barcode, barcode.removesuffix(".0")):
            marker = f"{prefix}-"
            if description.casefold().startswith(marker.casefold()):
                description = description[len(marker):].strip()
                break

    description = re.sub(r"^[A-Z]*\d[A-Z0-9]*\s*-\s*", "", description, flags=re.I)
    description = re.sub(r"^\d{3,5}\s*(?=[A-Za-zÀ-ÿ])", "", description)
    for _ in range(5):
        parts = re.split(r"\s*-\s*", description)
        if len(parts) < 2:
            break
        tail = fold(parts[-1])
        if tail not in DESCRIPTION_COLOR_WORDS and tail not in DESCRIPTION_SIZE_WORDS:
            break
        description = " - ".join(parts[:-1]).strip(" -")
    return clean(description)


def canonical_tokens(value: Any) -> set[str]:
    text = fold(value)
    phrase_aliases = {
        "tank top": "debardeur",
        "crew neck": "col rond",
        "v neck": "col v",
        "no sleeve": "sans manche",
        "short sleeved": "manche courte",
        "short sleeve": "manche courte",
        "long sleeved": "manche longue",
        "long sleeve": "manche longue",
        "wide band": "bande large",
    }
    for source, destination in phrase_aliases.items():
        text = text.replace(source, destination)
    return {
        TOKEN_ALIASES.get(word, word)
        for word in text.split()
        if word not in MATCH_STOPWORDS and not word.isdigit()
    }


def read_alsamah_rows() -> list[dict[str, Any]]:
    if not ALSAMAH_BOOK.exists():
        return []
    ws = load_workbook(ALSAMAH_BOOK, read_only=True, data_only=True).active
    rows = ws.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    result = []
    for row in rows:
        item = dict(zip(headers, row))
        if clean(item.get("DESCRIPTION")):
            result.append(item)
    return result


def best_existing_product(description: str, products: list[dict[str, Any]]) -> str | None:
    description_folded = fold(description)
    description_tokens = canonical_tokens(description)
    matches = []
    for product in products:
        name = product["name"]["fr"]
        name_tokens = canonical_tokens(name)
        overlap = len(description_tokens & name_tokens)
        coverage = overlap / max(1, min(len(description_tokens), len(name_tokens)))
        union = len(description_tokens | name_tokens)
        jaccard = overlap / max(1, union)
        sequence = SequenceMatcher(None, description_folded, fold(name)).ratio()
        score = 0.25 * sequence + 0.45 * coverage + 0.30 * jaccard
        matches.append((score, sequence, coverage, product["id"]))
    matches.sort(reverse=True)
    if not matches:
        return None
    best = matches[0]
    margin = best[0] - matches[1][0] if len(matches) > 1 else best[0]
    if (best[0] >= 0.82 and margin >= 0.035) or (best[1] >= 0.92 and best[2] >= 0.8):
        return best[3]
    return None


def rebuild_alsamah(existing_products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = read_alsamah_rows()
    existing_by_id = {product["id"]: product for product in existing_products}
    existing_by_name: dict[str, list[str]] = defaultdict(list)
    barcode_to_product: dict[str, str] = {}
    for product in existing_products:
        existing_by_name[fold(product["name"]["fr"])].append(product["id"])
        for variant in product.get("variants", []):
            barcode = clean(variant.get("barcode") or variant.get("id"))
            if barcode:
                barcode_to_product[barcode] = product["id"]

    assignments: list[str | None] = []
    normalized_names = [normalize_alsamah_description(item) for item in rows]
    for item, normalized_name in zip(rows, normalized_names):
        barcode = clean(item.get("Code Barre"))
        product_id = barcode_to_product.get(barcode)
        if product_id is None:
            normalized_id = slug(normalized_name)
            product_id = PRODUCT_ALIASES.get(normalized_id)
            if product_id not in existing_by_id:
                product_id = None
        if product_id is None:
            for candidate in (slug(clean(item.get("DESCRIPTION"))), slug(normalized_name)):
                if candidate in existing_by_id:
                    product_id = candidate
                    break
        if product_id is None:
            for candidate in (fold(clean(item.get("DESCRIPTION"))), fold(normalized_name)):
                ids = existing_by_name.get(candidate, [])
                if len(ids) == 1:
                    product_id = ids[0]
                    break
        assignments.append(product_id)

    learned: dict[str, Counter[str]] = defaultdict(Counter)
    for normalized_name, product_id in zip(normalized_names, assignments):
        if product_id:
            learned[slug(normalized_name)][product_id] += 1
    for index, product_id in enumerate(assignments):
        if product_id is not None:
            continue
        candidates = learned.get(slug(normalized_names[index]))
        if candidates:
            assignments[index] = candidates.most_common(1)[0][0]

    for index, product_id in enumerate(assignments):
        if product_id is not None:
            continue
        matched = best_existing_product(normalized_names[index], existing_products)
        assignments[index] = matched or slug(normalized_names[index])

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    names_by_product: dict[str, Counter[str]] = defaultdict(Counter)
    for item, normalized_name, product_id in zip(rows, normalized_names, assignments):
        assert product_id is not None
        grouped[product_id].append(item)
        names_by_product[product_id][normalized_name] += 1

    products = []
    for product_id, variants_source in grouped.items():
        existing = existing_by_id.get(product_id)
        display_name = existing["name"]["fr"] if existing else names_by_product[product_id].most_common(1)[0][0]
        variant_rows = []
        for item in variants_source:
            color_id = slug(color_key(item.get("Color")))
            size = clean(item.get("Size")) or "TU"
            barcode = clean(item.get("Code Barre")) or None
            quantity = max(0, int(number(item.get("QUANTITE")) or 0))
            variant_regular = number(item.get("PRIX DE VENTE"))
            variant_rows.append({
                "id": barcode or f"{product_id}-{slug(size)}-{color_id}",
                "colorId": color_id,
                "size": size,
                "stock": quantity,
                "barcode": barcode,
                "regularPrice": variant_regular,
                "price": discounted_price(variant_regular),
            })

        priced_variants = [variant for variant in variant_rows if variant["regularPrice"] is not None]
        representative = representative_variant(variant_rows)
        regular = representative["regularPrice"] if representative else None
        current_price = representative["price"] if representative else None

        preserved_colors = {
            color.get("id"): {
                key: color.get(key)
                for key in ("image", "imageKind", "lifestyleImage")
                if color.get(key) is not None
            }
            for color in (existing or {}).get("colors", [])
        }
        colors = []
        for color_name in dict.fromkeys(color_key(item.get("Color")) for item in variants_source):
            entry = color_entry(color_name)
            exact_asset = real_color_asset(product_id, entry["id"])
            if exact_asset:
                entry["image"], entry["imageKind"] = exact_asset
            elif preserved_colors.get(entry["id"], {}).get("image"):
                entry.update(preserved_colors[entry["id"]])
            colors.append(entry)

        category = category_from_name(display_name)
        stock = sum(variant["stock"] for variant in variant_rows)
        source_page = next((clean(item.get("MODELE")) for item in variants_source if clean(item.get("MODELE")).startswith("http")), None)
        product = {
            "id": product_id,
            "brand": "ALSAMAH",
            "category": category,
            "categoryName": {"femme": "Femme", "homme": "Homme", "enfants": "Enfants", "autres": "Collants & chaussettes"}[category],
            "subcategory": subcategory(category, display_name),
            "name": (existing or {}).get("name", {"fr": display_name, "ar": display_name, "en": display_name}),
            "short": (existing or {}).get("short", {
                "fr": f"{display_name}, selection ALSAMAH.",
                "ar": display_name,
                "en": f"{display_name}, an ALSAMAH essential.",
            }),
            "description": (existing or {}).get("description", {
                "fr": "Produit ALSAMAH disponible dans les tailles et couleurs indiquees.",
                "ar": display_name,
                "en": "ALSAMAH product available in the listed sizes and colors.",
            }),
            "regularPrice": regular,
            "price": current_price,
            "stock": stock,
            "sizes": sorted({variant["size"] for variant in variant_rows}),
            "colors": colors,
            "variants": variant_rows,
            "purchasable": any(variant["price"] is not None and variant["stock"] > 0 for variant in variant_rows),
            "imageStatus": "generated" if any(color.get("imageKind") == "generated" for color in colors) else "source" if any(color.get("image") for color in colors) else "missing",
            "fallbackImage": False,
            "sourcePage": source_page,
            "sourceWorkbook": ALSAMAH_BOOK.name,
            "variantCount": len(variant_rows),
            "missing": {"price": not priced_variants, "category": False, "image": not any(color.get("image") for color in colors)},
        }
        products.append(product)
    return products


def elko_base_name(description: str) -> str:
    return re.sub(r"\s*-\s*(XS|S|M|L|XL|XXL|2XL|3XL|4XL|5XL|6XL)\s*-\s*(White|Black)\s*$", "", clean(description), flags=re.I)


def read_elko() -> list[dict[str, Any]]:
    if not ELKO_BOOK.exists():
        return []
    workbook = load_workbook(ELKO_BOOK, read_only=True, data_only=True)
    products: list[dict[str, Any]] = []
    used: set[str] = set()
    for ws in workbook.worksheets:
        rows = ws.iter_rows(values_only=True)
        headers = [clean(value) for value in next(rows)]
        groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            item = dict(zip(headers, row))
            name = elko_base_name(item.get("DESCRIPTION"))
            if name:
                groups[name].append(item)
        for name, variants in groups.items():
            category = "homme" if ws.title.upper() == "MEN" else "femme"
            product_id = f"elko-{slug(name)}"
            base = product_id
            suffix = 2
            while product_id in used:
                product_id = f"{base}-{suffix}"
                suffix += 1
            used.add(product_id)
            stock = sum(max(0, number(item.get("QUANTITE")) or 0) for item in variants)
            colors = []
            for color in dict.fromkeys(color_key(item.get("Color")) for item in variants):
                colors.append(color_entry(color))
            variant_rows = []
            variant_ids: Counter[str] = Counter()
            for item in variants:
                qty = max(0, int(number(item.get("QUANTITE")) or 0))
                variant_regular = number(item.get("PRIX DE VENTE"))
                barcode = clean(item.get("Code Barre")) or None
                base_variant_id = barcode or f"{product_id}-{slug(item.get('Size'))}-{slug(item.get('Color'))}"
                variant_ids[base_variant_id] += 1
                variant_id = base_variant_id if variant_ids[base_variant_id] == 1 else f"{base_variant_id}-{variant_ids[base_variant_id]}"
                variant_rows.append({
                    "id": variant_id,
                    "colorId": slug(color_key(item.get("Color"))),
                    "size": clean(item.get("Size")) or "TU",
                    "stock": qty,
                    "barcode": barcode,
                    "regularPrice": variant_regular,
                    "price": discounted_price(variant_regular),
                })
            priced_variants = [variant for variant in variant_rows if variant["regularPrice"] is not None]
            representative = representative_variant(variant_rows)
            regular = representative["regularPrice"] if representative else None
            current_price = representative["price"] if representative else None
            products.append({
                "id": product_id,
                "brand": "ELKO",
                "category": category,
                "categoryName": "Homme" if category == "homme" else "Femme",
                "subcategory": subcategory(category, name),
                "name": {"fr": name, "ar": name, "en": name},
                "short": {
                    "fr": f"{name}, selection ELKO.",
                    "ar": name,
                    "en": f"{name}, an ELKO essential.",
                },
                "description": {
                    "fr": "Produit ELKO disponible dans les tailles et couleurs indiquees.",
                    "ar": name,
                    "en": "ELKO product available in the listed sizes and colors.",
                },
                "regularPrice": regular,
                "price": current_price,
                "stock": stock,
                "sizes": sorted({variant["size"] for variant in variant_rows}),
                "colors": colors,
                "variants": variant_rows,
                "purchasable": any(variant["price"] is not None and variant["stock"] > 0 for variant in variant_rows),
                "imageStatus": "missing",
                "fallbackImage": False,
                "sourcePage": None,
                "sourceWorkbook": ELKO_BOOK.name,
                "variantCount": len(variant_rows),
                "missing": {"price": not priced_variants, "category": False, "image": True},
            })
    return products


def enrich_existing(product: dict[str, Any], source_groups: dict[str, list[dict[str, Any]]]) -> None:
    group = source_groups.get(product["id"]) or source_groups.get(slug(product["name"]["fr"]))
    if group:
        variants = []
        for item in group:
            qty = max(0, int(number(item.get("QUANTITE")) or 0))
            variant_regular = number(item.get("PRIX DE VENTE"))
            variants.append({
                "id": clean(item.get("Code Barre")) or f"{product['id']}-{slug(item.get('Size'))}-{slug(item.get('Color'))}",
                "colorId": slug(color_key(item.get("Color"))),
                "size": clean(item.get("Size")) or "TU",
                "stock": qty,
                "barcode": clean(item.get("Code Barre")) or None,
                "regularPrice": variant_regular,
                "price": discounted_price(variant_regular),
            })
        priced_variants = [variant for variant in variants if variant["regularPrice"] is not None]
        representative = representative_variant(variants)
        regular = representative["regularPrice"] if representative else product.get("regularPrice")
        product["variants"] = variants
        product["stock"] = sum(variant["stock"] for variant in variants)
        product["variantCount"] = len(variants)
        product["regularPrice"] = regular
        product["price"] = representative["price"] if representative else sale_price(regular, product["stock"])
        product["sizes"] = sorted({variant["size"] for variant in variants}) or product.get("sizes", [])
        existing_colors = {color["id"] for color in product.get("colors", [])}
        for item in group:
            color = color_key(item.get("Color"))
            color_id = slug(color)
            if color_id not in existing_colors:
                product.setdefault("colors", []).append(color_entry(color))
                existing_colors.add(color_id)

    if product.get("brand") == "ALSAMAH":
        category = category_from_name(product["name"]["fr"])
        product["category"] = category
        product["categoryName"] = {"femme": "Femme", "homme": "Homme", "enfants": "Enfants", "autres": "Collants & chaussettes"}[category]
        product.setdefault("missing", {})["category"] = False

    product["subcategory"] = subcategory(product["category"], product["name"]["fr"])

    has_real_image = False
    for color in product.get("colors", []):
        if color.get("image"):
            if color.get("imageKind") != "fallback":
                has_real_image = True
            continue
        color["image"] = None
        color["imageKind"] = "missing"
        color.pop("fallbackImage", None)
    if not product.get("colors"):
        product["colors"] = [color_entry("non renseignee")]

    has_any_image = any(color.get("image") for color in product["colors"])
    if product.get("imageStatus") == "missing" or not has_real_image:
        product["imageStatus"] = "source" if has_real_image else "missing"
    product["fallbackImage"] = False
    product.setdefault("missing", {})["image"] = not has_any_image
    product["purchasable"] = bool(product.get("price") is not None and product.get("stock", 0) > 0)


def normalize_prices(products: list[dict[str, Any]]) -> None:
    for product in products:
        variants = product.get("variants", [])
        for variant in variants:
            regular = variant.get("regularPrice")
            if regular is None:
                regular = product.get("regularPrice")
                variant["regularPrice"] = regular
            variant["price"] = discounted_price(float(regular)) if regular is not None else None
        priced_variants = [variant for variant in variants if variant.get("regularPrice") is not None]
        representative = representative_variant(variants)
        if representative:
            product["regularPrice"] = representative["regularPrice"]
            product["price"] = representative["price"]
        else:
            regular = product.get("regularPrice")
            product["price"] = sale_price(float(regular), float(product.get("stock") or 0)) if regular is not None else None
        product.setdefault("missing", {})["price"] = product.get("regularPrice") is None
        product["purchasable"] = any(variant.get("price") is not None and variant.get("stock", 0) > 0 for variant in variants) if variants else bool(product["price"] is not None and product.get("stock", 0) > 0)


def main() -> None:
    existing_products = json.loads(CATALOG.read_text(encoding="utf-8"))
    existing_by_id = {product["id"]: product for product in existing_products}
    products = rebuild_alsamah([product for product in existing_products if product.get("brand") == "ALSAMAH"])

    for product in read_elko():
        existing = existing_by_id.get(product["id"])
        images_by_color = {
            color.get("id"): {
                key: color.get(key)
                for key in ("image", "imageKind", "lifestyleImage", "fallbackImage")
                if color.get(key) is not None
            }
            for color in (existing or {}).get("colors", [])
        }
        for color in product.get("colors", []):
            exact_asset = real_color_asset(product["id"], color["id"])
            preserved = images_by_color.get(color.get("id"))
            if exact_asset:
                color["image"], color["imageKind"] = exact_asset
            elif preserved and preserved.get("image"):
                color.update(preserved)
        has_any_image = any(color.get("image") for color in product.get("colors", []))
        product["imageStatus"] = (
            "generated"
            if any(color.get("imageKind") == "generated" for color in product.get("colors", []))
            else "source"
            if has_any_image
            else "missing"
        )
        product["fallbackImage"] = False
        product.setdefault("missing", {})["image"] = not has_any_image
        products.append(product)

    normalize_prices(products)

    products.sort(key=lambda product: (
        product.get("brand") != "ALSAMAH",
        not product.get("purchasable"),
        -float(product.get("stock") or 0),
        product["name"]["fr"].casefold(),
    ))
    CATALOG.write_text(json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "products": len(products),
        "with_images": sum(bool(product.get("colors")) and any(color.get("image") for color in product["colors"]) for product in products),
        "purchasable": sum(bool(product.get("purchasable")) for product in products),
        "missing_images": sum(product.get("missing", {}).get("image") for product in products),
        "elko": sum(product.get("brand") == "ELKO" for product in products),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
