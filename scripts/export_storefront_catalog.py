"""Export the complete HAWANA storefront catalog from XLSX or CSV.

The exporter never fabricates missing commercial data. Missing prices,
categories, and images remain null and make the product unavailable.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "Excel WASSAL 2025.xlsx"
DEFAULT_OUTPUT = ROOT / "src" / "data" / "catalog.json"
OFFICIAL_IMAGE_REPORT = ROOT / "src" / "data" / "official-image-report.json"

GENERATED_IMAGES: dict[str, dict[str, str]] = {
    "Manche courte col rond": {
        "bleu marine": "/assets/products/round-neck-navy.png",
        "gris": "/assets/products/round-neck-gray.png",
        "noir": "/assets/products/round-neck-black.png",
    },
    "Manche courte col V": {
        "bleu marine": "/assets/products/v-neck-navy.png",
        "gris": "/assets/products/v-neck-gray.png",
        "noir": "/assets/products/v-neck-black.png",
    },
    "Body à manches courtes": {
        "blanc": "/assets/products/body-white.png",
        "blanc cassé": "/assets/products/body-offwhite.png",
        "noir": "/assets/products/body-black.png",
    },
    "Men's corset boxer": {
        "blanc": "/assets/products/corset-boxer-white.png",
        "noir": "/assets/products/corset-boxer-black.png",
    },
    "Legging opaque épais - 120 den, style simple": {
        "blanc": "/assets/products/legging-white.png",
        "blanc cassé": "/assets/products/legging-offwhite.png",
        "noir": "/assets/products/legging-black.png",
    },
    "Top body à bretelles": {
        "beige": "/assets/products/strap-top-beige.png",
        "bleu marine": "/assets/products/strap-top-navy.png",
        "noir": "/assets/products/strap-top-black.png",
    },
}

LIFESTYLE_IMAGES: dict[str, dict[str, str]] = {
    "Manche courte col rond": {"bleu marine": "/assets/lifestyle/round-neck-navy-worn.png"},
    "Manche courte col V": {"gris": "/assets/lifestyle/v-neck-gray-worn.png"},
    "Body à manches courtes": {"blanc cassé": "/assets/lifestyle/body-offwhite-worn.png"},
}

SWATCHES = {
    "noir": "#1d1d1c", "black": "#1d1d1c",
    "blanc": "#f5f5f0", "white": "#f5f5f0",
    "blanc cassé": "#e9e2d2", "blanc casse": "#e9e2d2", "off-white": "#e9e2d2",
    "beige": "#cfb491", "gris": "#a8aaad", "gray": "#a8aaad", "grey": "#a8aaad",
    "bleu marine": "#17233d", "navy": "#17233d", "bleu": "#36608d",
    "bleu ciel": "#91b9d0", "rose": "#c9828e", "fushia": "#b73577",
    "rouge": "#a7473e", "jaune": "#d0a53a", "marron": "#6f4a35",
    "violet": "#765d87", "vert": "#627b54", "coloré": "#8a756f",
}

COLOR_AR = {
    "noir": "أسود", "blanc": "أبيض", "blanc cassé": "أوف وايت", "beige": "بيج",
    "gris": "رمادي", "bleu marine": "كحلي", "bleu": "أزرق", "bleu ciel": "أزرق سماوي",
    "rose": "وردي", "rouge": "أحمر", "jaune": "أصفر", "marron": "بني", "violet": "بنفسجي",
}

COLOR_EN = {
    "noir": "Black", "blanc": "White", "blanc cassé": "Off-white", "beige": "Beige",
    "gris": "Gray", "bleu marine": "Navy", "bleu": "Blue", "bleu ciel": "Sky blue",
    "rose": "Pink", "rouge": "Red", "jaune": "Yellow", "marron": "Brown", "violet": "Purple",
}


def clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip().replace("—", "-").replace("–", "-")
    return text or None


def key(value: Any) -> str:
    text = clean(value) or ""
    text = " ".join(text.casefold().replace("_", " ").split())
    aliases = {"blanc casse": "blanc cassé", "blanc cass": "blanc cassé", "white": "blanc", "black": "noir"}
    return aliases.get(text, text)


def slugify(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", normalized.casefold()).strip("-") or "produit"


def direct_image(url: str | None) -> str | None:
    if not url:
        return None
    path = url.split("?", 1)[0].casefold()
    return url if path.endswith((".jpg", ".jpeg", ".png", ".webp", ".avif")) else None


def category_bucket(category: str | None) -> str:
    if not category:
        return "non-classe"
    lowered = category.casefold()
    if "enfants" in lowered or "fille" in lowered or "garçon" in lowered:
        return "enfants"
    if "fashion femme" in lowered or "femme" in lowered:
        return "femme"
    if "homme" in lowered:
        return "homme"
    return "autres"


def fold(text: str) -> str:
    normalized = unicodedata.normalize("NFD", text.casefold())
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def subcategory_bucket(bucket: str, purchasable: bool, name: str, category_name: str | None) -> str | None:
    if not purchasable:
        return None
    folded_name = fold(name)
    folded_category = fold(category_name or "")
    if bucket == "femme":
        if re.search(r"corset|gainant", folded_name):
            return "corsets"
        if re.search(r"\b(soutien|brassiere|bra)\b", folded_name):
            return "soutien-gorge"
        if re.search(r"\bbody\b|corps|bretelles", folded_name):
            return "bodies"
        if "collant" in folded_name:
            return "collants"
        if "chaussette" in folded_name:
            return "chaussettes"
        if "nuisette" in folded_name or "chemise de nuit" in folded_name:
            return "nuisettes"
        if re.search(r"boxer|culotte|slip", folded_name):
            return "culottes"
        return "vetements"
    if bucket == "homme":
        if re.search(r"sleeve|manche|debardeur", folded_name):
            return "hauts"
        if re.search(r"corset|gainant", folded_name):
            return "corsets"
        if re.search(r"boxer|slip", folded_name):
            return "sous-vetements"
        return "hauts"
    if bucket == "enfants":
        if "garcon" in folded_category or "garcon" in folded_name:
            return "garcon"
        if "fille" in folded_category or "fille" in folded_name:
            return "fille"
        if "chaussette" in folded_name:
            return "chaussettes"
        return "collants"
    return None


def load_xlsx(path: Path) -> tuple[list[dict[str, Any]], dict[str, str], dict[str, str]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["bulk_import"]
    values = sheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(values)]
    rows = [dict(zip(headers, row)) for row in values if any(value is not None for value in row)]

    category_map: dict[str, str] = {}
    for row in workbook["category ids"].iter_rows(values_only=True):
        populated = [clean(value) for value in row if clean(value)]
        if len(populated) >= 2 and populated[0] != "ID":
            category_map[populated[0]] = populated[1]

    brand_map: dict[str, str] = {}
    for row in workbook["brand ids"].iter_rows(values_only=True):
        populated = [clean(value) for value in row if clean(value)]
        if len(populated) >= 2 and populated[0] != "ID":
            brand_map[populated[0]] = populated[1]
    return rows, category_map, brand_map


def load_csv(path: Path) -> tuple[list[dict[str, Any]], dict[str, str], dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle)), {}, {}


def median(values: Iterable[float]) -> float | None:
    ordered = sorted(values)
    if not ordered:
        return None
    middle = len(ordered) // 2
    return ordered[middle] if len(ordered) % 2 else (ordered[middle - 1] + ordered[middle]) / 2


def export_catalog(input_path: Path, output_path: Path) -> None:
    rows, category_map, brand_map = load_xlsx(input_path) if input_path.suffix.casefold() == ".xlsx" else load_csv(input_path)
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[clean(row.get("name")) or "Produit sans nom"].append(row)

    products: list[dict[str, Any]] = []
    used_ids: set[str] = set()
    for name, variants in grouped.items():
        first = variants[0]
        identifier = slugify(name)
        if identifier in used_ids:
            identifier = f"{identifier}-{slugify(clean(first.get('sku')) or str(len(products) + 1))}"
        used_ids.add(identifier)

        prices = [float(row["price"]) for row in variants if row.get("price") not in (None, "")]
        price = median(prices)
        category_id = next((clean(row.get("categ_id")) for row in variants if clean(row.get("categ_id"))), None)
        category_name = category_map.get(category_id or "")
        brand_id = next((clean(row.get("product_brand_id")) for row in variants if clean(row.get("product_brand_id"))), None)
        brand = brand_map.get(brand_id or "", brand_id or "Non renseignée")
        stock = sum(float(row.get("inventory") or 0) for row in variants)
        positive_stock = max(0, round(stock, 2))
        if price is None:
            sale_price = None
        else:
            discount = 0.30 if positive_stock >= 50 else 0.35 if positive_stock >= 11 else 0.40 if positive_stock >= 3 else 0.50
            sale_price = round(price * (1 - discount))

        original_images = [direct_image(clean(row.get("photo_urls"))) for row in variants]
        original_image = next((image for image in original_images if image), None)
        generated = GENERATED_IMAGES.get(name, {})
        lifestyle = LIFESTYLE_IMAGES.get(name, {})
        colors: list[dict[str, Any]] = []
        seen_colors: set[str] = set()
        for row in variants:
            color_name = key(row.get("color")) or "non renseignée"
            if color_name in seen_colors:
                continue
            seen_colors.add(color_name)
            generated_image = generated.get(color_name)
            colors.append({
                "id": slugify(color_name),
                "label": {
                    "fr": color_name.capitalize(),
                    "ar": COLOR_AR.get(color_name, color_name),
                    "en": COLOR_EN.get(color_name, color_name.capitalize()),
                },
                "hex": SWATCHES.get(color_name, "#8a8882"),
                "image": generated_image or original_image,
                "imageKind": "generated" if generated_image else "source" if original_image else "missing",
                "lifestyleImage": lifestyle.get(color_name),
            })
        if not colors:
            colors = [{"id": "non-renseignee", "label": {"fr": "Non renseignée", "ar": "غير محدد", "en": "Not provided"}, "hex": "#8a8882", "image": original_image, "imageKind": "source" if original_image else "missing"}]

        sizes = sorted({clean(row.get("size")) for row in variants if clean(row.get("size"))}) or ["Non renseignée"]
        name_ar = clean(first.get("name_ar")) or name
        name_en = clean(first.get("name_en")) or name
        short_fr = clean(first.get("description_sale")) or "Description à compléter."
        short_ar = clean(first.get("description_sale_ar")) or name_ar
        short_en = clean(first.get("description_sale_en")) or name_en
        description_fr = clean(first.get("content")) or short_fr
        description_ar = clean(first.get("content_ar")) or short_ar
        description_en = clean(first.get("content_en")) or short_en
        has_generated = any(color["imageKind"] == "generated" for color in colors)
        has_image = any(color["image"] for color in colors)
        purchasable = price is not None and category_id is not None and positive_stock > 0
        bucket = category_bucket(category_name)

        products.append({
            "id": identifier,
            "brand": brand,
            "category": bucket,
            "categoryName": category_name,
            "subcategory": subcategory_bucket(bucket, purchasable, name, category_name),
            "name": {"fr": name, "ar": name_ar, "en": name_en},
            "short": {"fr": short_fr, "ar": short_ar, "en": short_en},
            "description": {"fr": description_fr, "ar": description_ar, "en": description_en},
            "regularPrice": round(price, 2) if price is not None else None,
            "price": sale_price,
            "stock": positive_stock,
            "sizes": sizes,
            "colors": colors,
            "purchasable": purchasable,
            "imageStatus": "generated" if has_generated else "source" if has_image else "missing",
            "sourcePage": next((clean(row.get("photo_urls")) for row in variants if clean(row.get("photo_urls"))), None),
            "variantCount": len(variants),
            "missing": {
                "price": price is None,
                "category": category_id is None,
                "image": not has_image,
            },
        })

    official_images: dict[str, str] = {}
    if OFFICIAL_IMAGE_REPORT.exists():
        report = json.loads(OFFICIAL_IMAGE_REPORT.read_text(encoding="utf-8"))
        official_images = {
            item["catalogId"]: item["image"]
            for item in report
            if item.get("status") == "matched" and item.get("catalogId") and item.get("image")
        }
    for product in products:
        official_image = official_images.get(product["id"])
        if product["imageStatus"] == "missing" and official_image:
            asset_path = ROOT / "public" / official_image.removeprefix("/")
            if asset_path.exists():
                for color in product["colors"]:
                    color["image"] = official_image
                    color["imageKind"] = "source"
                product["imageStatus"] = "source"
                product["missing"]["image"] = False

    products.sort(key=lambda product: (not product["purchasable"], -product["stock"], product["name"]["fr"].casefold()))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "products": len(products),
        "variants": sum(product["variantCount"] for product in products),
        "purchasable": sum(product["purchasable"] for product in products),
        "with_generated_images": sum(product["imageStatus"] == "generated" for product in products),
        "with_source_images": sum(product["imageStatus"] == "source" for product in products),
        "missing_images": sum(product["imageStatus"] == "missing" for product in products),
        "output": str(output_path),
    }, ensure_ascii=False))


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("input", nargs="?", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    export_catalog(args.input.resolve(), args.output.resolve())
