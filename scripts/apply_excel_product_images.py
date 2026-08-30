from __future__ import annotations

import json
import mimetypes
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any
from urllib.error import URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public"
CATALOG = ROOT / "src" / "data" / "catalog.json"
ALSAMAH_BOOK = ROOT / "alsamah 18-08-2026 (1).xlsx"
ELKO_BOOK = ROOT / "Les articles disponibles ELKO depot 17-08-2026.xlsx"
REPORT = ROOT / "src" / "data" / "excel-product-image-report.json"


def clean(value: Any) -> str:
    text = str(value or "").replace("\ufffd", "").strip()
    return re.sub(r"\s+", " ", text).strip(" -")


def fold(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text.casefold()).strip()


def slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "-", fold(value)).strip("-") or "produit"


def color_key(value: Any) -> str:
    value = fold(value)
    aliases = {
        "white": "blanc", "black": "noir", "gray": "gris", "grey": "gris",
        "navy": "bleu marine", "pink": "rose", "blue": "bleu",
        "griis": "gris", "lanc": "blanc", "soman": "saumon", "somon": "saumon",
        "bordou": "bordeaux", "blanc cass": "blanc casse", "blanc cassee": "blanc casse",
    }
    return aliases.get(value, value or "non renseignee")


def elko_base_name(description: Any) -> str:
    return re.sub(r"\s*-\s*(XS|S|M|L|XL|XXL|2XL|3XL|4XL|5XL|6XL)\s*-\s*(White|Black)\s*$", "", clean(description), flags=re.I)


def extension_from_bytes(data: bytes, fallback: str = ".png") -> str:
    if data.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if data.startswith(b"RIFF") and data[8:12] == b"WEBP":
        return ".webp"
    return fallback


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for index in range(2, 1000):
        candidate = path.with_name(f"{stem}-{index}{suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Unable to allocate a filename for {path}")


def public_url(path: Path) -> str:
    return "/" + path.relative_to(PUBLIC).as_posix()


def direct_image_url(url: str) -> str | None:
    return url if re.search(r"\.(png|jpe?g|webp|avif)(\?|$)", url, re.I) else None


def fetch_url(url: str, timeout: int = 35) -> bytes | None:
    try:
      request = Request(url, headers={"User-Agent": "Mozilla/5.0 HAWANA launch image audit"})
      with urlopen(request, timeout=timeout) as response:
          return response.read()
    except (OSError, URLError, ValueError):
      return None


def scrape_product_image(url: str) -> str | None:
    data = fetch_url(url)
    if not data:
        return None
    html = data.decode("utf-8", errors="ignore")
    patterns = [
        r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']',
        r'<meta[^>]+name=["\']twitter:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'(https://cdn\.shopify\.com/[^"\']+\.(?:jpg|jpeg|png|webp)(?:\?[^"\']*)?)',
        r'(https://www\.alsamah\.com/cdn/shop/[^"\']+\.(?:jpg|jpeg|png|webp)(?:\?[^"\']*)?)',
    ]
    for pattern in patterns:
        match = re.search(pattern, html, re.I)
        if match:
            image = match.group(1)
            if image.startswith("//"):
                return "https:" + image
            if image.startswith("/"):
                return "https://www.alsamah.com" + image
            return image
    return None


def save_remote_image(url: str, destination: Path) -> str | None:
    data = fetch_url(url)
    if not data:
        return None
    content_type = mimetypes.guess_type(url.split("?", 1)[0])[0]
    suffix = mimetypes.guess_extension(content_type or "") or extension_from_bytes(data)
    if suffix == ".jpe":
        suffix = ".jpg"
    path = unique_path(destination.with_suffix(suffix))
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)
    return public_url(path)


def extract_embedded_images(path: Path, brand: str) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    workbook = load_workbook(path, read_only=False, data_only=True)
    extracted: dict[str, dict[str, str]] = defaultdict(dict)
    for ws in workbook.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        headers = [clean(value) for value in rows[0]]
        desc_i = headers.index("DESCRIPTION") if "DESCRIPTION" in headers else None
        color_i = headers.index("Color") if "Color" in headers else None
        if desc_i is None:
            continue
        row_data = {index + 1: row for index, row in enumerate(rows[1:], start=1)}
        for image in getattr(ws, "_images", []):
            row_index = image.anchor._from.row + 1
            row = row_data.get(row_index)
            if not row:
                continue
            raw_name = row[desc_i]
            name = elko_base_name(raw_name) if brand == "ELKO" else clean(raw_name)
            product_id = f"elko-{slug(name)}" if brand == "ELKO" else slug(name)
            color_id = slug(color_key(row[color_i] if color_i is not None and color_i < len(row) else "default"))
            data = image._data()
            suffix = extension_from_bytes(data)
            destination = PUBLIC / "assets" / "excel-products" / brand.lower() / f"{product_id}-{color_id}{suffix}"
            destination.parent.mkdir(parents=True, exist_ok=True)
            if not destination.exists():
                destination.write_bytes(data)
            extracted[product_id].setdefault(color_id, public_url(destination))
            extracted[product_id].setdefault("__first__", public_url(destination))
    return extracted


def workbook_urls(path: Path) -> dict[str, list[str]]:
    if not path.exists():
        return {}
    urls: dict[str, list[str]] = defaultdict(list)
    workbook = load_workbook(path, read_only=True, data_only=True)
    for ws in workbook.worksheets:
        rows = ws.iter_rows(values_only=True)
        headers = [clean(value) for value in next(rows)]
        desc_i = headers.index("DESCRIPTION") if "DESCRIPTION" in headers else None
        model_i = headers.index("MODELE") if "MODELE" in headers else None
        if desc_i is None or model_i is None:
            continue
        for row in rows:
            name = clean(row[desc_i])
            url = clean(row[model_i])
            if not name or not url.startswith("http"):
                continue
            urls[slug(name)].append(url)
    return {key: list(dict.fromkeys(value)) for key, value in urls.items()}


def apply_images() -> None:
    catalog = json.loads(CATALOG.read_text(encoding="utf-8"))
    embedded = extract_embedded_images(ALSAMAH_BOOK, "ALSAMAH")
    embedded.update(extract_embedded_images(ELKO_BOOK, "ELKO"))
    urls = workbook_urls(ALSAMAH_BOOK)
    report = []

    for product in catalog:
        product_id = product["id"]
        source_map = embedded.get(product_id) or embedded.get(slug(product["name"]["fr"])) or {}
        if not source_map and product.get("brand") == "ALSAMAH":
            for source_url in urls.get(product_id, []):
                image_url = direct_image_url(source_url) or scrape_product_image(source_url)
                if not image_url:
                    continue
                saved = save_remote_image(image_url, PUBLIC / "assets" / "excel-products" / "alsamah-url" / product_id)
                if saved:
                    source_map = {"__first__": saved}
                    break

        if source_map:
            for color in product.get("colors", []):
                exact_source = source_map.get(color["id"])
                if color.get("imageKind") == "generated" and not (product.get("brand") == "ELKO" and exact_source):
                    continue
                image = exact_source or (None if product.get("brand") == "ELKO" else source_map.get("__first__"))
                if image:
                    color["image"] = image
                    color["imageKind"] = "source"
                    color.pop("fallbackImage", None)
            product["imageStatus"] = "generated" if any(color.get("imageKind") == "generated" for color in product.get("colors", [])) else "source"
            product["fallbackImage"] = False
            product.setdefault("missing", {})["image"] = False
            report.append({"id": product_id, "name": product["name"]["fr"], "brand": product["brand"], "status": "mapped", "image": source_map.get("__first__")})
            continue

        for color in product.get("colors", []):
            if color.get("imageKind") == "fallback":
                color["image"] = None
                color["imageKind"] = "missing"
                color.pop("fallbackImage", None)
        product["imageStatus"] = "generated" if any(color.get("imageKind") == "generated" for color in product.get("colors", [])) else "source" if any(color.get("image") for color in product.get("colors", [])) else "missing"
        product["fallbackImage"] = False
        product.setdefault("missing", {})["image"] = not any(color.get("image") for color in product.get("colors", []))
        if product["missing"]["image"]:
            report.append({"id": product_id, "name": product["name"]["fr"], "brand": product["brand"], "status": "no_excel_image"})

    CATALOG.write_text(json.dumps(catalog, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "mapped": sum(item["status"] == "mapped" for item in report),
        "no_excel_image": sum(item["status"] == "no_excel_image" for item in report),
        "report": str(REPORT.relative_to(ROOT)),
    }, ensure_ascii=False))


if __name__ == "__main__":
    apply_images()
