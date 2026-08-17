# -*- coding: utf-8 -*-
import openpyxl, collections, statistics, json, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = 'c:/Users/Asus/Desktop/alsamah/Excel WASSAL 2025.xlsx'
OUT = 'c:/Users/Asus/Desktop/alsamah/LIQUIDATION_WASSAL_2025_MAD.xlsx'
SCRATCH = 'C:/Users/Asus/AppData/Local/Temp/claude/c--Users-Asus-Desktop-alsamah/f2a15547-3e7a-4a87-980e-4a67d676ab73/scratchpad'

wb = openpyxl.load_workbook(SRC, data_only=True)
rows = list(wb['bulk_import'].iter_rows(values_only=True))
hdr = [h for h in rows[0] if h]
data = []
for i, r in enumerate(rows[1:], start=2):
    if any(x is not None for x in r[:20]):
        d = dict(zip(hdr, r))
        d['_row'] = i
        data.append(d)

cat = {}
for r in wb['category ids'].iter_rows(min_row=4, values_only=True):
    if r[1] is not None:
        cat[str(r[1]).strip()] = str(r[2]).strip()
brands = {}
for r in wb['brand ids'].iter_rows(min_row=4, values_only=True):
    if r[2] is not None:
        brands[str(r[2]).strip()] = str(r[3]).strip()

# --- backfill price from priced sibling variants of the same parent product
byname = collections.defaultdict(list)
for d in data:
    byname[d['name']].append(d)
for n, g in byname.items():
    ps = [d['price'] for d in g if d.get('price') is not None]
    med = statistics.median(ps) if ps else None
    for d in g:
        if d.get('price') is not None:
            d['price_source'] = 'fichier'
        elif med:
            d['price'] = med
            d['price_source'] = 'estime (variante soeur)'
        else:
            d['price_source'] = 'MANQUANT'

for d in data:
    d['_inv'] = d.get('inventory') or 0
    d['_p'] = d.get('price') or 0
    d['_val'] = d['_inv'] * d['_p']
    cid = str(d.get('categ_id')).strip() if d.get('categ_id') is not None else None
    d['_cat'] = cat.get(cid, '') if cid else ''
    d['_cat_top'] = d['_cat'].split('/')[0].strip() if d['_cat'] else 'Non classe'
    d['_brand'] = brands.get(str(d.get('product_brand_id')).strip(), '')


def issues(d):
    it = []
    if d['price_source'] == 'MANQUANT':
        it.append('Prix manquant')
    if not d.get('categ_id'):
        it.append('Categorie manquante')
    if not d.get('photo_urls'):
        it.append('Photo manquante')
    if not d.get('content'):
        it.append('Fiche produit vide')
    if d['_inv'] <= 0:
        it.append('Stock <= 0')
    return it


for d in data:
    d['_iss'] = issues(d)


def tier(d):
    if d['price_source'] == 'MANQUANT' or not d.get('categ_id'):
        return ('X', 'Bloque - a completer', 0.0)
    i = d['_inv']
    if i <= 0:
        return ('X', 'Bloque - stock nul', 0.0)
    if i <= 2:
        return ('D', 'Lot / bundle mixte', 0.50)
    if i <= 10:
        return ('C', 'Destockage', 0.40)
    if i <= 49:
        return ('B', 'Remise ciblee', 0.35)
    return ('A', 'Vente flash volume', 0.30)


for d in data:
    t, lbl, dis = tier(d)
    d['_tier'], d['_tlabel'], d['_dis'] = t, lbl, dis
    d['_pliq'] = round(d['_p'] * (1 - dis)) if dis else 0
    d['_rev'] = d['_pliq'] * d['_inv']

TOT_U = sum(d['_inv'] for d in data)
TOT_V = sum(d['_val'] for d in data)
SELL = [d for d in data if d['_tier'] != 'X']
BLOCK = [d for d in data if d['_tier'] == 'X']
REV = sum(d['_rev'] for d in SELL)
V_SELL = sum(d['_val'] for d in SELL)
V_BLOCK = sum(d['_val'] for d in BLOCK)

H = PatternFill('solid', fgColor='1F2937')
HF = Font(color='FFFFFF', bold=True, size=10)
TIERFILL = {'A': 'DCFCE7', 'B': 'DBEAFE', 'C': 'FEF3C7', 'D': 'FFE4E6', 'X': 'F3F4F6'}
thin = Side(style='thin', color='D1D5DB')
BD = Border(left=thin, right=thin, top=thin, bottom=thin)
out = openpyxl.Workbook()
out.remove(out.active)


def sheet(name, headers, rws, widths=None, money_cols=(), fills=None):
    ws = out.create_sheet(name)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = H
        cell.font = HF
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='center')
    ws.row_dimensions[1].height = 30
    for i, r in enumerate(rws, start=2):
        ws.append(r)
        if fills:
            f = fills(rws[i - 2])
            if f:
                for c in range(1, len(headers) + 1):
                    ws.cell(i, c).fill = PatternFill('solid', fgColor=f)
        for c in range(1, len(headers) + 1):
            ws.cell(i, c).border = BD
    for c in money_cols:
        for i in range(2, len(rws) + 2):
            ws.cell(i, c).number_format = '#,##0'
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(headers)), len(rws) + 1)
    return ws


# ---------- 1. RESUME ----------
ws = out.create_sheet('1. Resume')
ws.column_dimensions['A'].width = 48
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 52


def line(a, b='', c='', bold=False, fill=None, size=11):
    ws.append([a, b, c])
    r = ws.max_row
    ws.cell(r, 1).font = Font(bold=bold, size=size)
    ws.cell(r, 2).font = Font(bold=True, size=size)
    ws.cell(r, 2).number_format = '#,##0'
    ws.cell(r, 3).font = Font(size=9, color='6B7280')
    ws.cell(r, 3).alignment = Alignment(wrap_text=True)
    if fill:
        for k in range(1, 4):
            ws.cell(r, k).fill = PatternFill('solid', fgColor=fill)
    return r


line('PLAN DE LIQUIDATION - STOCK WASSAL 2025', bold=True, size=16, fill='1F2937')
for k in range(1, 4):
    ws.cell(1, k).font = Font(bold=True, size=14, color='FFFFFF')
line('')
line('ETAT DU STOCK', bold=True, size=12, fill='E5E7EB')
line('Lignes (variantes SKU)', len(data), '1 ligne = 1 couleur + 1 taille')
line('Produits parents distincts', len(byname))
line('Unites physiques en stock', TOT_U)
line('Valeur stock au prix catalogue (MAD)', round(TOT_V), 'Inclut les prix estimes par variante soeur')
line('')
line('CAPACITE DE VENTE IMMEDIATE', bold=True, size=12, fill='E5E7EB')
line('Unites vendables maintenant', sum(d['_inv'] for d in SELL),
     '%.0f%% du stock' % (100 * sum(d['_inv'] for d in SELL) / TOT_U))
line('Valeur vendable (catalogue)', round(V_SELL))
line('Recette estimee apres remises', round(REV),
     'Remise moyenne ponderee %.0f%%' % (100 * (1 - REV / V_SELL)))
line('Cash a laisser sur la table', round(V_SELL - REV))
line('')
line('STOCK BLOQUE - ACTION REQUISE', bold=True, size=12, fill='FEE2E2')
line('Lignes bloquees', len(BLOCK), 'Ni prix ni categorie -> impossible a lister/vendre')
line('Unites bloquees', sum(d['_inv'] for d in BLOCK),
     '%.0f%% du stock physique dort' % (100 * sum(d['_inv'] for d in BLOCK) / TOT_U))
line('Valeur potentielle bloquee (est.)', round(V_BLOCK), 'Fortement sous-estimee: la plupart nont aucun prix de reference')
line('')
line('QUALITE DES DONNEES', bold=True, size=12, fill='E5E7EB')
qual = [
    ('Lignes sans prix', sum(1 for d in data if d['price_source'] == 'MANQUANT')),
    ('Lignes sans categorie', sum(1 for d in data if not d.get('categ_id'))),
    ('Lignes sans photo', sum(1 for d in data if not d.get('photo_urls'))),
    ('Lignes sans fiche produit (content)', sum(1 for d in data if not d.get('content'))),
    ('Lignes stock <= 0', sum(1 for d in data if d['_inv'] <= 0)),
    ('Codes-barres dupliques', sum(1 for k, v in collections.Counter(str(d.get('barcode')) for d in data).items() if v > 1)),
    ('SKU variantes dupliques', sum(1 for k, v in collections.Counter(str(d.get('sku_variant')) for d in data).items() if v > 1)),
]
for lbl, n in qual:
    line(lbl, n, ('%.0f%% des lignes' % (100 * n / len(data))) if n else 'OK')
line('')
line('STRATEGIE PAR PALIER', bold=True, size=12, fill='E5E7EB')
tiers = collections.defaultdict(lambda: [0, 0, 0, 0])
for d in data:
    a = tiers[(d['_tier'], d['_tlabel'])]
    a[0] += 1
    a[1] += d['_inv']
    a[2] += d['_val']
    a[3] += d['_rev']
ws.append(['Palier', 'Unites', 'Recette estimee (MAD)'])
r = ws.max_row
for k in range(1, 4):
    ws.cell(r, k).fill = H
    ws.cell(r, k).font = HF
DISMAP = {'A': 30, 'B': 35, 'C': 40, 'D': 50, 'X': 0}
for (t, lbl), v in sorted(tiers.items()):
    dis = ('-%d%%' % DISMAP[t]) if DISMAP[t] else 'n/a'
    ws.append(['%s - %s (%s)' % (t, lbl, dis), v[1], v[3]])
    r = ws.max_row
    for k in range(1, 4):
        ws.cell(r, k).fill = PatternFill('solid', fgColor=TIERFILL[t])
        ws.cell(r, k).border = BD
    ws.cell(r, 2).number_format = '#,##0'
    ws.cell(r, 3).number_format = '#,##0'

# ---------- 2. PLAN DE LIQUIDATION ----------
H2 = ['Palier', 'Action', 'Produit', 'Couleur', 'Taille', 'SKU variante', 'Code-barres', 'Categorie',
      'Stock', 'Prix catalogue', 'Source prix', 'Remise', 'Prix liquidation', 'Recette estimee', 'Problemes']
rws = []
for d in sorted(data, key=lambda x: (x['_tier'], -x['_rev'], -x['_val'])):
    rws.append([d['_tier'], d['_tlabel'], d['name'], d.get('color') or '', str(d.get('size') or ''),
                d.get('sku_variant') or '', str(d.get('barcode') or ''), d['_cat'] or 'NON CLASSE',
                d['_inv'], d['_p'] or '', d['price_source'],
                ('-%d%%' % int(d['_dis'] * 100)) if d['_dis'] else '',
                d['_pliq'] or '', d['_rev'] or '', ' | '.join(d['_iss'])])
sheet('2. Plan de liquidation', H2, rws,
      widths=[7, 22, 42, 16, 10, 20, 16, 40, 8, 13, 20, 9, 15, 15, 45],
      money_cols=(9, 10, 13, 14),
      fills=lambda r: TIERFILL.get(r[0]))

# ---------- 3. PAR PRODUIT ----------
pag = collections.defaultdict(lambda: {'v': 0, 'u': 0, 'val': 0, 'rev': 0, 'cat': '', 'sizes': set(), 'colors': set(), 'p': [], 'blk': 0})
for d in data:
    a = pag[d['name']]
    a['v'] += 1
    a['u'] += d['_inv']
    a['val'] += d['_val']
    a['rev'] += d['_rev']
    a['cat'] = a['cat'] or d['_cat']
    a['sizes'].add(str(d.get('size') or ''))
    a['colors'].add(str(d.get('color') or ''))
    if d['_p']:
        a['p'].append(d['_p'])
    if d['_tier'] == 'X':
        a['blk'] += 1
H3 = ['Produit', 'Categorie', 'Nb variantes', 'Dont bloquees', 'Unites', 'Couleurs', 'Tailles',
      'Prix moyen', 'Valeur catalogue', 'Recette estimee', 'Priorite']
rws = []
for n, a in sorted(pag.items(), key=lambda x: -x[1]['val']):
    pr = 'HAUTE' if a['val'] >= 15000 else 'MOYENNE' if a['val'] >= 4000 else 'BASSE'
    if a['blk'] == a['v']:
        pr = 'BLOQUE'
    rws.append([n, a['cat'] or 'NON CLASSE', a['v'], a['blk'], a['u'], len(a['colors']), len(a['sizes']),
                round(statistics.mean(a['p'])) if a['p'] else '', round(a['val']), round(a['rev']), pr])
sheet('3. Par produit', H3, rws, widths=[46, 38, 12, 13, 9, 10, 9, 12, 17, 17, 11], money_cols=(9, 10),
      fills=lambda r: {'HAUTE': 'DCFCE7', 'MOYENNE': 'DBEAFE', 'BASSE': 'FEF3C7', 'BLOQUE': 'FEE2E2'}.get(r[10]))

# ---------- 4. PAR CATEGORIE ----------
cg = collections.defaultdict(lambda: [0, 0, 0, 0])
for d in data:
    a = cg[(d['_cat_top'], d['_cat'] or 'NON CLASSE')]
    a[0] += 1
    a[1] += d['_inv']
    a[2] += d['_val']
    a[3] += d['_rev']
rws = [[k[0], k[1], v[0], v[1], round(v[2]), round(v[3])] for k, v in sorted(cg.items(), key=lambda x: -x[1][2])]
sheet('4. Par categorie', ['Univers', 'Categorie complete', 'Lignes', 'Unites', 'Valeur catalogue', 'Recette estimee'],
      rws, widths=[22, 58, 10, 10, 18, 18], money_cols=(5, 6))

# ---------- 5. LOTS ----------
lots = collections.defaultdict(lambda: [0, 0, 0])
for d in data:
    if d['_tier'] == 'D':
        a = lots[(d['_cat_top'], d['_cat'] or 'NON CLASSE')]
        a[0] += 1
        a[1] += d['_inv']
        a[2] += d['_val']
rws = []
for k, v in sorted(lots.items(), key=lambda x: -x[1][1]):
    nlots = max(1, v[1] // 5)
    rws.append([k[0], k[1], v[0], v[1], round(v[2]), nlots, round(v[2] * 0.5 / nlots)])
sheet('5. Lots a composer',
      ['Univers', 'Categorie', 'Lignes orphelines', 'Unites', 'Valeur catalogue', 'Nb lots de 5 pcs', 'Prix conseille / lot'],
      rws, widths=[22, 52, 17, 10, 18, 17, 20], money_cols=(5, 7))

# ---------- 6. A CORRIGER ----------
H6 = ['Ligne Excel', 'Produit', 'SKU variante', 'Taille', 'Couleur', 'Stock', 'Prix manquant',
      'Categorie manquante', 'Photo manquante', 'Fiche vide', 'Blocage total']
rws = []
for d in sorted(data, key=lambda x: -x['_inv']):
    if d['_iss']:
        rws.append([d['_row'], d['name'], d.get('sku_variant') or '', str(d.get('size') or ''), d.get('color') or '', d['_inv'],
                    'OUI' if 'Prix manquant' in d['_iss'] else '',
                    'OUI' if 'Categorie manquante' in d['_iss'] else '',
                    'OUI' if 'Photo manquante' in d['_iss'] else '',
                    'OUI' if 'Fiche produit vide' in d['_iss'] else '',
                    'OUI' if d['_tier'] == 'X' else ''])
sheet('6. A corriger', H6, rws, widths=[12, 44, 20, 10, 16, 8, 14, 20, 16, 12, 14], money_cols=(6,),
      fills=lambda r: 'FEE2E2' if r[10] == 'OUI' else 'FEF3C7')

out.save(OUT)
print('WROTE ' + OUT)

payload = {
    'totals': {'rows': len(data), 'parents': len(byname), 'units': TOT_U, 'value': round(TOT_V),
               'sell_units': sum(d['_inv'] for d in SELL), 'sell_value': round(V_SELL), 'revenue': round(REV),
               'block_rows': len(BLOCK), 'block_units': sum(d['_inv'] for d in BLOCK), 'block_value': round(V_BLOCK)},
    'quality': {'no_price': qual[0][1], 'no_cat': qual[1][1], 'no_photo': qual[2][1],
                'no_content': qual[3][1], 'zero_stock': qual[4][1], 'dup_sku': qual[6][1]},
    'tiers': [{'t': t, 'label': lbl, 'rows': v[0], 'units': v[1], 'value': round(v[2]), 'rev': round(v[3]),
               'dis': DISMAP[t]} for (t, lbl), v in sorted(tiers.items())],
    'cats': [{'name': k[0], 'full': k[1], 'rows': v[0], 'units': v[1], 'value': round(v[2]), 'rev': round(v[3])}
             for k, v in sorted(cg.items(), key=lambda x: -x[1][2])][:12],
    'products': [{'name': n, 'cat': a['cat'] or 'Non classe', 'var': a['v'], 'blk': a['blk'], 'units': a['u'],
                  'value': round(a['val']), 'rev': round(a['rev'])}
                 for n, a in sorted(pag.items(), key=lambda x: -x[1]['val'])][:40],
    'depth': dict(collections.Counter(
        ('0' if d['_inv'] <= 0 else '1-2' if d['_inv'] <= 2 else '3-5' if d['_inv'] <= 5
         else '6-10' if d['_inv'] <= 10 else '11-20' if d['_inv'] <= 20 else '21-50' if d['_inv'] <= 50 else '50+')
        for d in data)),
    'sizes': collections.Counter(str(d.get('size') or '') for d in data).most_common(10),
    'colors': collections.Counter(str(d.get('color') or '').lower() for d in data).most_common(10),
}
with io.open(SCRATCH + '/data.json', 'w', encoding='utf-8') as f:
    json.dump(payload, f, ensure_ascii=False)
print('JSON ok')
